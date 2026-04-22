import openpyxl
import re
from collections import defaultdict
import unicodedata

wb = openpyxl.load_workbook(r"c:\Users\Noutbuk savdosi\CRM\03.03.2026.xlsx", data_only=True)

# Use the latest sheet - "март 2026"
ws = wb['март 2026']

# Column 2 = фирма (client name)
# Column 28 = Ост на 31.03.2026 (latest balance)
FIRM_COL = 2
BALANCE_COL = 28
DATA_START_ROW = 4  # rows 1-3 are headers

def normalize_name(name):
    """Normalize client name for grouping similar names."""
    if not name or not isinstance(name, str):
        return None
    
    s = name.strip()
    # Lowercase
    s = s.lower()
    # Remove extra spaces
    s = re.sub(r'\s+', ' ', s)
    # Remove common prefixes
    for prefix in ['ооо ', 'оoo ', 'чп ', 'ип ', 'ао ', 'зао ', 'тоо ', 'мчж ', 'xk ', 'хк ']:
        if s.startswith(prefix):
            s = s[len(prefix):]
    s = s.strip()
    
    # Remove quotes
    s = s.replace('"', '').replace("'", '').replace('«', '').replace('»', '')
    s = s.strip()
    
    return s if s else None

# Collect raw data: client -> list of balances from all their rows
raw_clients = defaultdict(list)
all_rows_data = []

for row_idx in range(DATA_START_ROW, ws.max_row + 1):
    firm = ws.cell(row=row_idx, column=FIRM_COL).value
    balance = ws.cell(row=row_idx, column=BALANCE_COL).value
    
    if firm is None or not isinstance(firm, str) or firm.strip() == '':
        continue
    
    firm_clean = firm.strip()
    
    # Convert balance to number
    if balance is None:
        bal = 0
    elif isinstance(balance, (int, float)):
        bal = float(balance)
    elif isinstance(balance, str):
        # Try to parse as number
        try:
            bal = float(balance.replace(' ', '').replace(',', '.'))
        except ValueError:
            bal = 0
    else:
        bal = 0
    
    all_rows_data.append((firm_clean, bal, row_idx))

print(f"Total data rows with firm names: {len(all_rows_data)}")

# Group by normalized name
grouped = defaultdict(lambda: {'original_names': set(), 'total_balance': 0, 'row_count': 0})

for firm_clean, bal, row_idx in all_rows_data:
    norm = normalize_name(firm_clean)
    if norm is None:
        continue
    grouped[norm]['original_names'].add(firm_clean)
    grouped[norm]['total_balance'] += bal
    grouped[norm]['row_count'] += 1

print(f"Unique clients (after normalization): {len(grouped)}")

# Filter: only positive balances (debts)
debtors = {}
for norm_name, data in grouped.items():
    if data['total_balance'] > 0:
        # Use the most common original name as display name
        display_name = max(data['original_names'], key=len)
        debtors[display_name] = data['total_balance']

# Sort by debt descending
sorted_debtors = sorted(debtors.items(), key=lambda x: x[1], reverse=True)

total_debt = sum(d[1] for d in sorted_debtors)
num_debtors = len(sorted_debtors)

print("\n" + "="*70)
print(f"ОТЧЁТ О ДОЛГАХ КЛИЕНТОВ")
print(f"Источник: 03.03.2026.xlsx, лист 'март 2026'")
print(f"Колонка остатка: 'Ост на 31.03.2026' (колонка {BALANCE_COL})")
print("="*70)

print(f"\nОбщий долг: {total_debt:,.0f} сум".replace(',', ' '))
print(f"Количество должников: {num_debtors}")

print(f"\n{'='*70}")
print(f"ТОП-20 ДОЛЖНИКОВ:")
print(f"{'='*70}")
for i, (name, debt) in enumerate(sorted_debtors[:20], 1):
    debt_str = f"{debt:,.0f}".replace(',', ' ')
    print(f"{i:2d}. {name} — {debt_str} сум")

print(f"\n{'='*70}")
print(f"ПОЛНЫЙ СПИСОК ДОЛЖНИКОВ ({num_debtors}):")
print(f"{'='*70}")
for i, (name, debt) in enumerate(sorted_debtors, 1):
    debt_str = f"{debt:,.0f}".replace(',', ' ')
    print(f"{i:3d}. {name} — {debt_str} сум")

# Also show clients with multiple name variations
print(f"\n{'='*70}")
print(f"КЛИЕНТЫ С РАЗНЫМИ НАПИСАНИЯМИ:")
print(f"{'='*70}")
for norm_name, data in sorted(grouped.items()):
    if len(data['original_names']) > 1:
        print(f"  Объединены: {data['original_names']} -> сумма: {data['total_balance']:,.0f}".replace(',', ' '))
