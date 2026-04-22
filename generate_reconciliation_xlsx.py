"""
Генерация финального Excel-отчёта реконсилиэйшн v2.0
с NetDebt, PrepaySum, TakenOnCredit колонками.
Запускается отдельно от основного аудита.
"""
import sys, os, json, csv
sys.stdout.reconfigure(encoding='utf-8')

from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import psycopg2
import re

ROOT = Path(r"c:\Users\Noutbuk savdosi\CRM")
OUT  = ROOT / "mnt" / "data"

DB_URL = (
    "postgresql://crm_user:BChpe9Gb4dOeVQQxRYVkiLUgu4TsmWJo"
    "@dpg-d6bcdrt6ubrc73ch10dg-a.oregon-postgres.render.com/crm_db_okj8"
)
TOLERANCE = 1_000

# Коды
CREDIT_CODES = {"к", "н/к", "п/к", "пк/к", "ф", "фт", "фотих"}
PREPAY_CODES = {"пп"}

# ---- Нормализация ----
def normalize_name(s):
    if s is None: return None
    s = str(s).strip().lower().replace("ё", "е")
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip()
    return s if s else None

# ---- Загрузить Excel-данные из CSV (v2.0: с prepay/net_debt) ----
def load_excel_debts_from_csv():
    """Читаем orders_clean CSV и агрегируем с prepay/net_debt."""
    cleaned = OUT / "cleaned"
    per_client = defaultdict(lambda: {
        "remaining": 0.0,
        "prepay_sum": 0.0,
        "taken_on_credit_sum": 0.0,
        "row_count": 0,
    })

    for file_dir in cleaned.iterdir():
        if not file_dir.is_dir():
            continue
        for csv_file in file_dir.glob("orders_clean_*.csv"):
            with open(csv_file, encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    j_str = (row.get("j_str", "") or "").strip().lower()
                    cnorm = (row.get("client_norm", "") or "").strip()
                    if not cnorm:
                        continue

                    is_credit = j_str in CREDIT_CODES
                    per_client[cnorm]["row_count"] += 1

                    try:
                        remaining = float(row.get("remaining_debt", 0) or 0)
                    except:
                        remaining = 0.0

                    try:
                        prepay = float(row.get("prepay_sum", 0) or 0)
                    except:
                        prepay = 0.0

                    if is_credit and remaining > 0:
                        per_client[cnorm]["remaining"] += remaining

                    per_client[cnorm]["prepay_sum"] += prepay

                    if is_credit:
                        per_client[cnorm]["taken_on_credit_sum"] += max(0, remaining)

    result = {}
    for cn, data in per_client.items():
        net_debt = data["remaining"] - data["prepay_sum"]
        result[cn] = {
            "remaining": data["remaining"],
            "prepay_sum": data["prepay_sum"],
            "taken_on_credit_sum": data["taken_on_credit_sum"],
            "net_debt": net_debt,
            "row_count": data["row_count"],
        }
    return result

# ---- Загрузить данные CRM ----
def load_crm_debts():
    conn = psycopg2.connect(DB_URL, connect_timeout=30)
    cur  = conn.cursor()
    cur.execute("""
        SELECT c.company_name, c.id,
               SUM(d.amount - d.paid_amount) AS debt,
               COUNT(d.id) AS dcnt
        FROM deals d
        JOIN clients c ON d.client_id = c.id
        WHERE d.payment_status IN ('UNPAID','PARTIAL')
          AND d.status NOT IN ('CANCELED','REJECTED')
          AND d.is_archived = false
        GROUP BY c.id, c.company_name
        HAVING SUM(d.amount - d.paid_amount) > 0
        ORDER BY debt DESC
    """)
    crm = {}
    total = 0.0
    for row in cur.fetchall():
        name, cid, debt, dcnt = row
        norm = normalize_name(name)
        debt = float(debt)
        crm[norm] = {"id": str(cid), "company_name": name, "debt": debt,
                     "deal_count": dcnt}
        total += debt
    cur.close(); conn.close()
    return crm, total

# ---- Levenshtein ----
def levenshtein(a, b):
    if len(a)<len(b): a,b=b,a
    if not b: return len(a)
    row=list(range(len(b)+1))
    for ca in a:
        nr=[row[0]+1]
        for j,cb in enumerate(b):
            nr.append(min(row[j+1]+1, nr[j]+1, row[j]+(ca!=cb)))
        row=nr
    return row[-1]

# ---- Реконсилиэйшн ----
def reconcile(excel_agg, crm, total_crm):
    en = set(excel_agg); cn = set(crm)
    diffs = []
    for nm in sorted(en & cn):
        ea = excel_agg[nm]
        cd = crm[nm]["debt"]
        diff = cd - ea["remaining"]
        if abs(diff) >= TOLERANCE:
            diffs.append({
                "client_norm": nm,
                "company_name": crm[nm]["company_name"],
                "excel_remaining": ea["remaining"],
                "excel_prepay": ea["prepay_sum"],
                "excel_net_debt": max(0, ea["net_debt"]),
                "excel_credit_sum": ea["taken_on_credit_sum"],
                "crm_debt": cd, "diff": diff,
                "status": "CRM_HIGHER" if diff>0 else "EXCEL_HIGHER",
            })
    diffs.sort(key=lambda x: abs(x["diff"]), reverse=True)

    only_crm = [{"client_norm":n, "company_name":crm[n]["company_name"],
                 "crm_debt":crm[n]["debt"], "deal_count":crm[n]["deal_count"]}
                for n in sorted(cn-en)]
    only_excel = []
    for n in sorted(en-cn):
        ea = excel_agg[n]
        if ea["remaining"] > 0:
            only_excel.append({
                "client_norm": n,
                "excel_debt": ea["remaining"],
                "excel_net_debt": max(0, ea["net_debt"]),
                "excel_prepay": ea["prepay_sum"],
            })

    total_excel = sum(ea["remaining"] for ea in excel_agg.values() if ea["remaining"] > 0)
    total_excel_net = sum(max(0, ea["net_debt"]) for ea in excel_agg.values())
    total_prepay = sum(ea["prepay_sum"] for ea in excel_agg.values())

    # Fuzzy cross-check
    fuzzy_matches = []
    for cn_norm in crm:
        if cn_norm in excel_agg: continue
        for en_norm in excel_agg:
            if en_norm in {d["client_norm"] for d in diffs}: continue
            ml = max(len(cn_norm),len(en_norm),1)
            thr = 2 if ml<=10 else max(2,int(ml*0.10))
            dist = levenshtein(cn_norm, en_norm)
            if dist <= thr:
                fuzzy_matches.append({"crm_norm": cn_norm, "excel_norm": en_norm,
                                      "crm_name": crm[cn_norm]["company_name"],
                                      "crm_debt": crm[cn_norm]["debt"],
                                      "excel_debt": excel_agg[en_norm]["remaining"],
                                      "distance": dist})
            elif " ".join(sorted(cn_norm.split())) == " ".join(sorted(en_norm.split())):
                fuzzy_matches.append({"crm_norm": cn_norm, "excel_norm": en_norm,
                                      "crm_name": crm[cn_norm]["company_name"],
                                      "crm_debt": crm[cn_norm]["debt"],
                                      "excel_debt": excel_agg[en_norm]["remaining"],
                                      "distance": 0})

    return {
        "total_crm": total_crm,
        "total_excel": total_excel,
        "total_excel_net": total_excel_net,
        "total_prepay": total_prepay,
        "global_diff": total_crm - total_excel,
        "diffs": diffs,
        "only_crm": only_crm,
        "only_excel": only_excel,
        "fuzzy_matches": fuzzy_matches,
    }

# ---- Загрузить аномалии ----
def load_anomalies():
    rpts = OUT / "reports"
    csvs = sorted(rpts.glob("anomalies_*.csv"))
    if not csvs: return []
    with open(csvs[-1], encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))

# ---- Генерация XLSX ----
def make_xlsx(rec, anomalies, baseline, total_crm_now, out_path,
              crm_client_count=0, excel_debtor_count=0):

    RED  = PatternFill("solid", fgColor="FFCCCC")
    YEL  = PatternFill("solid", fgColor="FFFACD")
    GRN  = PatternFill("solid", fgColor="CCFFCC")
    BLUE = PatternFill("solid", fgColor="CCE0FF")
    ORG  = PatternFill("solid", fgColor="FFE0B2")
    HDR  = PatternFill("solid", fgColor="2F5496")
    HFNT = Font(color="FFFFFF", bold=True)
    BFNT = Font(bold=True)

    def hrow(ws, cols):
        ws.append(cols)
        for c in ws[ws.max_row]:
            c.fill=HDR; c.font=HFNT; c.alignment=Alignment(horizontal="center")

    def aw(ws, extra=4):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=4)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+extra, 70)

    wb = openpyxl.Workbook()

    # ===== SUMMARY =====
    ws = wb.active; ws.title = "summary"
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append(["АУДИТ EXCEL vs CRM v2.0 -- ПОЛНЫЙ РЕКОНСИЛИЭЙШН"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Сформировано: {now_str}"])
    ws.append([])

    bl_total = baseline.get("total_debt", 0)
    bl_ts    = baseline.get("timestamp", "?")
    diff_bl  = total_crm_now - bl_total

    rows = [
        ("Baseline долг (начало аудита)", f"{bl_total:,.2f}", bl_ts[:19]),
        ("Текущий долг CRM",              f"{total_crm_now:,.2f}", f"delta={diff_bl:+,.2f}"),
        ("Суммарный остаток Excel (remaining)", f"{rec['total_excel']:,.2f}", "debt-типы: к/н/к/п/к/ф/фт/фотих"),
        ("Суммарная предоплата Excel",    f"{rec['total_prepay']:,.2f}", "ПП из строк и payment slots"),
        ("Net debt Excel",                f"{rec['total_excel_net']:,.2f}", "remaining - prepay_sum"),
        ("Разница (CRM - Excel remaining)", f"{rec['global_diff']:+,.2f}", "< 0 = Excel выше CRM"),
        ("Клиентов с долгом в CRM",       str(crm_client_count), ""),
        ("Клиентов с долгом в Excel",     str(excel_debtor_count), ""),
        ("Клиентов только в CRM",         str(len(rec["only_crm"])), "не представлены в Excel"),
        ("Клиентов только в Excel",       str(len(rec["only_excel"])), "не представлены в CRM"),
        ("Клиентов с расхождением > 1000", str(len(rec["diffs"])), ""),
        ("Fuzzy-кандидатов для слияния",  str(len(rec["fuzzy_matches"])), "д.б. проверены вручную"),
        ("Всего аномалий в Excel",        str(len(anomalies)), ""),
        ("Baseline OK?",
         "YES" if abs(diff_bl) < TOLERANCE else f"DELTA {diff_bl:+,.2f}!",
         "долг CRM не изменился аудитом"),
    ]
    hrow(ws, ["Метрика", "Значение (сум)", "Примечание"])
    for r in rows:
        ws.append(list(r))
    aw(ws)

    # ===== DIFFS =====
    ws2 = wb.create_sheet("diffs")
    hrow(ws2, ["#","Клиент (norm)","Имя в CRM",
               "Excel Remaining","Excel Prepay","Excel NetDebt","Excel CreditSum",
               "Долг CRM","Разница (CRM-Excel)","Статус","Вероятная причина"])
    CAUSES = {
        "CRM_HIGHER":   "Оплата зафиксирована в Excel, но не проведена в CRM?",
        "EXCEL_HIGHER": "Сделка в CRM уже оплачена или архивирована; в Excel старый остаток",
    }
    for i, d in enumerate(rec["diffs"], 1):
        ws2.append([i, d["client_norm"], d["company_name"],
                    round(d["excel_remaining"],2), round(d["excel_prepay"],2),
                    round(d["excel_net_debt"],2), round(d.get("excel_credit_sum",0),2),
                    round(d["crm_debt"],2), round(d["diff"],2),
                    d["status"], CAUSES.get(d["status"],"--")])
        fill = RED if d["status"]=="CRM_HIGHER" else YEL
        for c in ws2[ws2.max_row]: c.fill = fill
    aw(ws2)

    # ===== TOP-20 =====
    ws3 = wb.create_sheet("top20_diffs")
    hrow(ws3, ["#","Клиент","Excel Remaining","Excel NetDebt","Долг CRM","Разница","Статус"])
    for i, d in enumerate(rec["diffs"][:20], 1):
        ws3.append([i, d["company_name"],
                    round(d["excel_remaining"],2), round(d["excel_net_debt"],2),
                    round(d["crm_debt"],2), round(d["diff"],2), d["status"]])
        for c in ws3[ws3.max_row]:
            c.fill = RED if d["diff"]>0 else YEL
    aw(ws3)

    # ===== ONLY CRM =====
    ws4 = wb.create_sheet("clients_only_in_CRM")
    ws4.append(["Клиенты присутствующие в CRM, но ОТСУТСТВУЮЩИЕ в Excel"])
    ws4["A1"].font = BFNT
    ws4.append([f"Итого: {len(rec['only_crm'])} | "
                f"Долг: {sum(c['crm_debt'] for c in rec['only_crm']):,.0f} сум"])
    ws4.append([])
    hrow(ws4, ["#","Клиент (norm)","Имя в CRM","Долг CRM","Сделок","Рекомендация"])
    for i,c in enumerate(rec["only_crm"],1):
        ws4.append([i, c["client_norm"], c["company_name"],
                    round(c["crm_debt"],2), c["deal_count"],
                    "Проверить: возможно под другим именем в Excel или архивировать"])
        for cell in ws4[ws4.max_row]: cell.fill = BLUE
    aw(ws4)

    # ===== ONLY EXCEL =====
    ws5 = wb.create_sheet("clients_only_in_Excel")
    ws5.append(["Клиенты присутствующие в Excel, но ОТСУТСТВУЮЩИЕ в CRM"])
    ws5["A1"].font = BFNT
    ws5.append([f"Итого: {len(rec['only_excel'])} | "
                f"Долг: {sum(c['excel_debt'] for c in rec['only_excel']):,.0f} сум"])
    ws5.append([])
    hrow(ws5, ["#","Клиент (norm)","Долг Excel","NetDebt","Prepay","Рекомендация"])
    for i,c in enumerate(rec["only_excel"],1):
        ws5.append([i, c["client_norm"], round(c["excel_debt"],2),
                    round(c.get("excel_net_debt",0),2),
                    round(c.get("excel_prepay",0),2),
                    "Создать в CRM или проверить нормализацию имени"])
        for cell in ws5[ws5.max_row]: cell.fill = GRN
    aw(ws5)

    # ===== FUZZY MATCHES =====
    ws6 = wb.create_sheet("fuzzy_merge_candidates")
    ws6.append(["Кандидаты на объединение имён (Левенштейн + word-sort)"])
    ws6["A1"].font = BFNT
    ws6.append(["ВАЖНО: Объединение вручную! Не объединять автоматически."])
    ws6.append([])
    hrow(ws6, ["#","Имя CRM (norm)","Имя Excel (norm)","Долг CRM","Долг Excel","Расстояние","Действие"])
    for i,m in enumerate(rec["fuzzy_matches"],1):
        ws6.append([i, m["crm_norm"], m["excel_norm"],
                    round(m["crm_debt"],2), round(m["excel_debt"],2),
                    m["distance"], "Проверить вручную"])
        for cell in ws6[ws6.max_row]: cell.fill = ORG
    aw(ws6)

    # ===== ANOMALIES =====
    ws7 = wb.create_sheet("anomalies")
    ws7.append([f"Аномалии в Excel -- всего {len(anomalies)}"])
    ws7["A1"].font = BFNT
    ws7.append([])
    if anomalies:
        anom_types = Counter(a.get("type","?") for a in anomalies)
        ws7.append(["Тип аномалии", "Кол-во", "Рекомендация"])
        RECS = {
            "non_numeric_balance": "=VALUE(SUBSTITUTE(ячейка,\" \",\"\")) -- удалить пробелы",
            "negative_balance": "Переплата -- проверить в CRM, создать refund payment",
            "missing_date_with_balance": "Вставить дату вручную",
            "ambiguous_payment_cell": "Ячейка суммы содержит текст -- разнести по колонкам",
        }
        for t, cnt in sorted(anom_types.items(), key=lambda x:-x[1]):
            ws7.append([t, cnt, RECS.get(t, "Проверить вручную")])
        ws7.append([])
        hrow(ws7, list(anomalies[0].keys()))
        for a in anomalies[:500]:
            ws7.append([str(a.get(k,"")) for k in anomalies[0].keys()])
    aw(ws7)

    # ===== PLAYBOOK =====
    ws8 = wb.create_sheet("playbook")
    ws8.append(["ПОШАГОВЫЙ ПЛЕЙБУК ДЛЯ РЕГУЛЯРНОГО ЗАПУСКА АУДИТА v2.0"])
    ws8["A1"].font = Font(bold=True, size=13)
    steps = [
        ("", ""),
        ("ШАГ 1", "Получить baseline: SELECT SUM(amount-paid_amount) FROM deals WHERE payment_status IN ('UNPAID','PARTIAL') AND is_archived=false"),
        ("ШАГ 2", "Создать бэкап Excel (автоматически через python excel_audit_full.py)"),
        ("ШАГ 3", "Запустить: python -X utf8 excel_audit_full.py (dry-run, только чтение)"),
        ("ШАГ 4", "Открыть excel_vs_crm_reconciliation.xlsx -> лист 'diffs' -> разобрать ТОП-20"),
        ("ШАГ 5", "Проверить prepay_sum и net_debt -- если клиент имеет prepay, net_debt может быть 0 или отрицательный"),
        ("ШАГ 6", "Для CRM_HIGHER: проверить payments в БД, возможно нужен sync-payments.ts"),
        ("ШАГ 7", "Для EXCEL_HIGHER: проверить, не пора ли архивировать старые PAID сделки"),
        ("ШАГ 8", "Лист 'clients_only_in_CRM': проверить каждого клиента"),
        ("ШАГ 9", "Лист 'fuzzy_merge_candidates': word-sort + Левенштейн -- вручную"),
        ("ШАГ 10", "После ручных правок -- повторить шаги 1-4, убедиться что baseline совпадает"),
        ("", ""),
        ("SQL DEBT CHECK", "SELECT c.company_name, SUM(d.amount-d.paid_amount) as debt FROM deals d JOIN clients c ON d.client_id=c.id WHERE d.payment_status IN ('UNPAID','PARTIAL') AND d.is_archived=false GROUP BY c.company_name ORDER BY debt DESC"),
        ("SQL OVERPAID", "SELECT id,amount,paid_amount FROM deals WHERE paid_amount > amount"),
        ("SQL OLD UNPAID", "SELECT id,amount,paid_amount,created_at FROM deals WHERE payment_status='UNPAID' AND created_at < '2025-01-01'"),
        ("", ""),
        ("МАППИНГ 2026", "Янв: J=10,AA=27 | Фев: J=10,AB=28 | Мар: J=10,AB=28"),
        ("МАППИНГ 2025", "Все месяцы: J=10, AA=27"),
        ("МАППИНГ 2024", "Янв: M=13,AF=32 | Фев-Окт: L=12,AC=29 | Ноя-Дек: J=10,AA=27"),
        ("", ""),
        ("CREDIT CODES", "к, н/к, п/к, пк/к, ф, фт, фотих"),
        ("PREPAY CODES", "пп, ПП"),
        ("NET DEBT", "net_debt = remaining - prepay_sum"),
        ("", ""),
        ("ПЕРИОДИЧНОСТЬ", "Запускать после каждого обновления Excel-файла"),
    ]
    hrow(ws8, ["Шаг/Команда","Описание/SQL"])
    for row in steps:
        ws8.append(list(row))
    aw(ws8)

    wb.save(str(out_path))
    print(f"Reconciliation XLSX saved: {out_path}")
    return out_path


# ---- MAIN ----
if __name__ == "__main__":
    print("=== GENERATING RECONCILIATION XLSX v2.0 ===")

    # Load baseline
    bl_path = OUT / "baseline_debt.json"
    with open(bl_path, encoding="utf-8") as f:
        baseline_data = json.load(f)
    print(f"Baseline: {baseline_data['total_debt']:,.2f} ({baseline_data['timestamp'][:19]})")

    # Load Excel data from existing CSVs
    print("Loading Excel debts from CSV (v2.0 with prepay/net_debt)...")
    excel_agg = load_excel_debts_from_csv()
    total_excel = sum(ea["remaining"] for ea in excel_agg.values() if ea["remaining"] > 0)
    total_prepay = sum(ea["prepay_sum"] for ea in excel_agg.values())
    total_net = sum(max(0, ea["net_debt"]) for ea in excel_agg.values())
    debtors_excel = len([ea for ea in excel_agg.values() if ea["remaining"] > 0])
    print(f"  Excel remaining: {debtors_excel} clients, total {total_excel:,.2f}")
    print(f"  Excel prepay:    total {total_prepay:,.2f}")
    print(f"  Excel net_debt:  total {total_net:,.2f}")

    # Load CRM data
    print("Loading CRM debts...")
    crm_clients, total_crm = load_crm_debts()
    print(f"  CRM: {len(crm_clients)} clients, total {total_crm:,.2f}")

    # Reconcile
    print("Reconciling...")
    rec = reconcile(excel_agg, crm_clients, total_crm)
    print(f"  Global diff (CRM-Excel): {rec['global_diff']:+,.2f}")
    print(f"  Diffs: {len(rec['diffs'])}, only_crm: {len(rec['only_crm'])}, only_excel: {len(rec['only_excel'])}")
    print(f"  Fuzzy candidates: {len(rec['fuzzy_matches'])}")

    print("\nTop-7 diffs:")
    print(f"  {'Client':<30} {'Remaining':>13} {'Prepay':>10} {'NetDebt':>12} {'CRM':>13} {'Diff':>13} {'Status'}")
    print("  " + "-"*100)
    for d in rec["diffs"][:7]:
        nm = d["company_name"][:28]
        print(f"  {nm:<30} {d['excel_remaining']:>13,.0f} {d['excel_prepay']:>10,.0f} "
              f"{d['excel_net_debt']:>12,.0f} {d['crm_debt']:>13,.0f} "
              f"{d['diff']:>13,.0f}  {d['status']}")

    # Load anomalies
    anomalies = load_anomalies()
    print(f"\nAnomalies loaded: {len(anomalies)}")

    # Verify baseline
    diff_bl = total_crm - baseline_data["total_debt"]
    if abs(diff_bl) < 1_000:
        print(f"\nBASELINE VERIFICATION: OK (diff={diff_bl:+,.2f})")
    else:
        print(f"\nBASELINE VERIFICATION: DELTA = {diff_bl:+,.2f} sum")

    # Generate XLSX
    out_path = OUT / "reports" / "excel_vs_crm_reconciliation.xlsx"
    make_xlsx(rec, anomalies, baseline_data, total_crm, out_path,
              crm_client_count=len(crm_clients),
              excel_debtor_count=debtors_excel)

    # Save fuzzy suggestions JSON
    fuzzy_path = OUT / "reports" / "fuzzy_suggestions_final.json"
    with open(fuzzy_path, "w", encoding="utf-8") as f:
        json.dump(rec["fuzzy_matches"], f, ensure_ascii=False, indent=2)
    print(f"Fuzzy suggestions: {fuzzy_path}")

    print("\n=== DONE ===")
    print(f"  Excel remaining : {total_excel:>20,.2f} sum")
    print(f"  Excel prepay    : {total_prepay:>20,.2f} sum")
    print(f"  Excel net_debt  : {total_net:>20,.2f} sum")
    print(f"  CRM total       : {total_crm:>20,.2f} sum")
    print(f"  Global diff     : {rec['global_diff']:>20,.2f} sum")
    print(f"  Baseline OK     : {'YES' if abs(diff_bl)<1000 else 'CHECK NEEDED'}")
    print(f"  Report          : {out_path}")
