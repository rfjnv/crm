"""
Аудит долгов v3: правильная логика выбора листа.
1. Для каждого клиента найти ПОСЛЕДНИЙ лист где он присутствует (любая строка)
2. Из этого листа взять ТОЛЬКО строки где J IN ('к', 'н/к', 'п/к', 'ф')
3. Суммировать колонку остатка по этим строкам
"""
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\Noutbuk savdosi\CRM\frontend\05.03.2026.xlsx', data_only=True)

sheets_config = [
    {'name': 'январь 2026',  'balance_col': 27},
    {'name': 'февраль 2026', 'balance_col': 28},
    {'name': 'март 2026',    'balance_col': 28},
]

CLIENT_COL = 2
J_COL = 10
DATA_START = 4
DEBT_TYPES = {'к', 'н/к', 'п/к', 'ф'}

def normalize(name):
    if not name:
        return None
    s = str(name).strip().lower().replace('ё', 'е')
    s = ' '.join(s.split())
    return s if s else None

# ============================================================
# Шаг 1: Собираем ВСЕ строки по каждому листу (без фильтра J)
# Для каждого клиента определяем последний лист присутствия
# ============================================================

# client_all_rows[norm] = {sheet_idx: [rows...]}
client_all_rows = {}

for sheet_idx, cfg in enumerate(sheets_config):
    ws = wb[cfg['name']]
    bal_col = cfg['balance_col']

    for row in range(DATA_START, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=CLIENT_COL).value
        norm = normalize(raw_name)
        if not norm:
            continue

        j_val = ws.cell(row=row, column=J_COL).value
        j_str = str(j_val).strip().lower() if j_val is not None else ''

        bal_val = ws.cell(row=row, column=bal_col).value
        try:
            balance = float(bal_val) if bal_val is not None else None
        except (ValueError, TypeError):
            balance = None

        if norm not in client_all_rows:
            client_all_rows[norm] = {}
        if sheet_idx not in client_all_rows[norm]:
            client_all_rows[norm][sheet_idx] = []

        client_all_rows[norm][sheet_idx].append({
            'raw_name': str(raw_name).strip(),
            'balance': balance,
            'j_type': j_str,
            'row': row,
        })

# ============================================================
# Шаг 2: Для каждого клиента — берём ПОСЛЕДНИЙ лист, потом фильтруем по J
# ============================================================

results = []

for norm, sheets_data in client_all_rows.items():
    # Последний лист где клиент присутствует
    last_sheet_idx = max(sheets_data.keys())
    rows = sheets_data[last_sheet_idx]

    # Имя клиента из первой строки
    raw_name = rows[0]['raw_name']
    sheet_name = sheets_config[last_sheet_idx]['name']

    # Фильтруем по J типам
    debt_rows = [r for r in rows if r['j_type'] in DEBT_TYPES and r['balance'] is not None]

    if not debt_rows:
        continue

    total_debt = sum(r['balance'] for r in debt_rows)
    row_count = len(debt_rows)

    # Разбивка по J
    j_breakdown = {}
    for r in debt_rows:
        jt = r['j_type']
        if jt not in j_breakdown:
            j_breakdown[jt] = {'sum': 0, 'count': 0}
        j_breakdown[jt]['sum'] += r['balance']
        j_breakdown[jt]['count'] += 1

    results.append({
        'name': raw_name,
        'norm': norm,
        'debt': total_debt,
        'rows': row_count,
        'sheet': sheet_name,
        'j_breakdown': j_breakdown,
    })

# Фильтруем и сортируем
debtors = [r for r in results if r['debt'] > 0]
debtors.sort(key=lambda x: x['debt'], reverse=True)

def fmt(n):
    if n == int(n):
        return f"{int(n):,}".replace(',', ' ')
    return f"{n:,.2f}".replace(',', ' ')

total_debt = sum(d['debt'] for d in debtors)
total_rows = sum(d['rows'] for d in debtors)
num_debtors = len(debtors)
avg_debt = total_debt / num_debtors if num_debtors else 0

# ============================================================
# Вывод
# ============================================================

print("=" * 95)
print("  ФИНАНСОВЫЙ АУДИТ ДОЛГОВ v3")
print(f"  Источник: 05.03.2026.xlsx")
print(f"  Фильтр J: к, н/к, п/к, ф")
print(f"  Логика: последний лист ПРИСУТСТВИЯ клиента → фильтр по J → SUM(AB)")
print("=" * 95)

print(f"\n  Общая сумма долга:          {fmt(total_debt)} сум")
print(f"  Количество должников:       {num_debtors}")
print(f"  Количество строк с долгом:  {total_rows}")
print(f"  Средний долг клиента:       {fmt(avg_debt)} сум")
if debtors:
    print(f"  Максимальный долг:          {fmt(debtors[0]['debt'])} сум ({debtors[0]['name']})")
    print(f"  Минимальный долг:           {fmt(debtors[-1]['debt'])} сум ({debtors[-1]['name']})")

# Разбивка по типам
total_by_j = {}
for d in debtors:
    for jt, info in d['j_breakdown'].items():
        if jt not in total_by_j:
            total_by_j[jt] = {'sum': 0, 'count': 0}
        total_by_j[jt]['sum'] += info['sum']
        total_by_j[jt]['count'] += info['count']

j_labels = {
    'к': 'к    — заказ в долг',
    'н/к': 'н/к  — наличная, в долг',
    'п/к': 'п/к  — перечисление, в долг',
    'ф': 'ф    — Фотих (свой человек)',
}
print(f"\n  Разбивка по типам J:")
for jt in ['к', 'н/к', 'п/к', 'ф']:
    if jt in total_by_j:
        info = total_by_j[jt]
        print(f"    {j_labels[jt]:<40} {fmt(info['sum']):>20}  ({info['count']} строк)")

# Полный список
print(f"\n{'─' * 95}")
print(f"  ПОЛНЫЙ СПИСОК ДОЛЖНИКОВ ({num_debtors})")
print(f"{'─' * 95}")
print(f"{'#':>4}  {'Клиент':<35} {'Долг (сум)':>20} {'Строк':>6}  {'Типы J':<15} {'Лист'}")
print("─" * 95)

for i, d in enumerate(debtors, 1):
    name = d['name'][:33] + '..' if len(d['name']) > 35 else d['name']
    j_types = ', '.join(sorted(d['j_breakdown'].keys()))
    print(f"{i:>4}  {name:<35} {fmt(d['debt']):>20} {d['rows']:>6}  {j_types:<15} {d['sheet']}")

print("─" * 95)
print(f"{'':>4}  {'ИТОГО':<35} {fmt(total_debt):>20} {total_rows:>6}")
print("=" * 95)

# Отдельно Ф
f_clients = [d for d in debtors if 'ф' in d['j_breakdown']]
if f_clients:
    print(f"\n  Должники с типом 'ф' (Фотих):")
    ft = 0
    for d in f_clients:
        fi = d['j_breakdown']['ф']
        ft += fi['sum']
        print(f"    {d['name']:<35} {fmt(fi['sum']):>20}")
    print(f"    {'ИТОГО Ф':<35} {fmt(ft):>20}")

# Сравнение с ожидаемым
expected = 1_159_293_323
diff = total_debt - expected
print(f"\n  Ожидаемый итог (пользователь):  {fmt(expected)} сум")
print(f"  Мой расчёт:                     {fmt(total_debt)} сум")
print(f"  Разница:                        {fmt(diff)} сум")
