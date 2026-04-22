# -*- coding: utf-8 -*-
import openpyxl
import sys
import re
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

filepath = 'c:/Users/Noutbuk savdosi/CRM/03.03.2026.xlsx'
wb = openpyxl.load_workbook(filepath, data_only=True)

def normalize_name(name):
    if not name or not isinstance(name, str):
        return None
    s = name.strip().lower()
    skip_words = ['итого', 'всего', 'итог', 'total']
    for sw in skip_words:
        if sw in s:
            return None
    for prefix in ['ооо ', 'ooo ', 'ип ', 'чп ', 'ао ', 'тоо ', 'мчж ']:
        if s.startswith(prefix):
            s = s[len(prefix):]
    s = s.strip()
    s = re.sub(r'\s+', ' ', s)
    if not s:
        return None
    return s

CLIENT_COL = 2

# Analyze each sheet independently
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Find ALL columns with balance-related headers
    print(f"\n{'='*70}")
    print(f"ЛИСТ: {sheet_name}")
    print(f"Строк: {ws.max_row}, Столбцов: {ws.max_column}")
    print(f"\nВСЕ заголовки (строка 1):")
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            print(f"  Col{c}: {v}")

    print(f"\nВСЕ заголовки (строка 2):")
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v:
            print(f"  Col{c}: {v}")

    # Find closing balance column
    balance_col = None
    for c in range(1, ws.max_column + 1):
        h1 = ws.cell(row=1, column=c).value
        if h1 and isinstance(h1, str) and 'Ост на' in h1 and 'Ост-к' not in h1:
            balance_col = c

    if not balance_col:
        print("  НЕТ СТОЛБЦА ОСТАТКА")
        continue

    print(f"\nСтолбец закрывающего остатка: Col{balance_col}")

    # Count clients and compute totals
    client_balances = defaultdict(float)
    client_rows = defaultdict(list)
    total_rows = 0

    for r in range(4, ws.max_row + 1):
        client_raw = ws.cell(row=r, column=CLIENT_COL).value
        if not client_raw or not isinstance(client_raw, str):
            continue
        client_raw = client_raw.strip()
        if not client_raw:
            continue
        norm = normalize_name(client_raw)
        if not norm:
            continue

        total_rows += 1
        balance_val = ws.cell(row=r, column=balance_col).value
        if balance_val is not None:
            try:
                balance_val = float(balance_val)
            except (ValueError, TypeError):
                balance_val = 0
        else:
            balance_val = 0

        client_balances[norm] += balance_val
        client_rows[norm].append((r, balance_val, client_raw))

    positive = {k: v for k, v in client_balances.items() if v > 0}
    total_positive = sum(positive.values())

    print(f"\nВсего строк данных: {total_rows}")
    print(f"Уникальных клиентов: {len(client_balances)}")
    print(f"Клиентов с положительным остатком: {len(positive)}")
    print(f"СУММА положительных остатков: {total_positive:,.0f} сум")

    # Check: sum of ALL closing balances (raw)
    raw_sum = 0
    for r in range(4, ws.max_row + 1):
        v = ws.cell(row=r, column=balance_col).value
        if v is not None:
            try:
                raw_sum += float(v)
            except:
                pass
    print(f"Сумма ВСЕХ остатков (включая отрицательные): {raw_sum:,.0f} сум")

# Now check: what does CRM use? Maybe it looks at February sheet
# since it's the last COMPLETE month?
print(f"\n{'='*70}")
print("СРАВНЕНИЕ: берём февраль как последний завершённый месяц")
print(f"{'='*70}")

ws = wb['февраль 2026']
balance_col = 28  # Ост на 28.02.2026

client_balances_feb = defaultdict(float)
for r in range(4, ws.max_row + 1):
    client_raw = ws.cell(row=r, column=CLIENT_COL).value
    if not client_raw or not isinstance(client_raw, str):
        continue
    client_raw = client_raw.strip()
    if not client_raw:
        continue
    norm = normalize_name(client_raw)
    if not norm:
        continue

    balance_val = ws.cell(row=r, column=balance_col).value
    if balance_val is not None:
        try:
            balance_val = float(balance_val)
        except (ValueError, TypeError):
            balance_val = 0
    else:
        balance_val = 0

    client_balances_feb[norm] += balance_val

positive_feb = {k: v for k, v in client_balances_feb.items() if v > 0}
total_feb = sum(positive_feb.values())
print(f"Февраль - должников: {len(positive_feb)}, сумма: {total_feb:,.0f} сум")

# Also check if CRM might be looking at opening balance of March (=closing of Feb)
print(f"\n{'='*70}")
print("ПРОВЕРКА: Открывающий остаток марта (= закрывающий февраля)")
print(f"{'='*70}")

ws_mar = wb['март 2026']
opening_col = 3  # Ост-к на 01.03.2026

client_opening_mar = defaultdict(float)
for r in range(4, ws_mar.max_row + 1):
    client_raw = ws_mar.cell(row=r, column=CLIENT_COL).value
    if not client_raw or not isinstance(client_raw, str):
        continue
    client_raw = client_raw.strip()
    if not client_raw:
        continue
    norm = normalize_name(client_raw)
    if not norm:
        continue

    balance_val = ws_mar.cell(row=r, column=opening_col).value
    if balance_val is not None:
        try:
            balance_val = float(balance_val)
        except (ValueError, TypeError):
            balance_val = 0
    else:
        balance_val = 0

    client_opening_mar[norm] += balance_val

positive_opening = {k: v for k, v in client_opening_mar.items() if v > 0}
total_opening = sum(positive_opening.values())
print(f"Март (открытие) - должников: {len(positive_opening)}, сумма: {total_opening:,.0f} сум")

# Check if the CRM sums across ALL sheets
print(f"\n{'='*70}")
print("ПРОВЕРКА: CRM суммирует остатки всех листов?")
print(f"{'='*70}")

all_sum = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    balance_col = None
    for c in range(1, ws.max_column + 1):
        h1 = ws.cell(row=1, column=c).value
        if h1 and isinstance(h1, str) and 'Ост на' in h1 and 'Ост-к' not in h1:
            balance_col = c
    if not balance_col:
        continue

    sheet_sum = 0
    for r in range(4, ws.max_row + 1):
        v = ws.cell(row=r, column=balance_col).value
        if v is not None:
            try:
                fv = float(v)
                if fv > 0:
                    sheet_sum += fv
            except:
                pass
    print(f"  {sheet_name}: положительные остатки = {sheet_sum:,.0f} сум")
    all_sum += sheet_sum

print(f"  ИТОГО (все листы): {all_sum:,.0f} сум")

# Maybe CRM uses the last row's "итого" row?
print(f"\n{'='*70}")
print("ПРОВЕРКА: Есть ли строка ИТОГО в конце каждого листа?")
print(f"{'='*70}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n  Лист: {sheet_name}, последние 5 строк:")
    for r in range(max(4, ws.max_row - 4), ws.max_row + 1):
        row_data = []
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_data.append(f"Col{c}={v}")
        print(f"    Row {r}: {row_data}")
