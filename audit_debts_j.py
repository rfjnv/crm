"""
Финансовый аудит долгов — с учётом колонки J (нкп).
Типы долга: "к", "н/к", "п/к", "ф"
Все эти типы означают что заказ может иметь задолженность.
Колонка AB (остаток) — источник истины.
"""
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\Noutbuk savdosi\CRM\03.03.2026.xlsx', data_only=True)

sheets_config = [
    {'name': 'январь 2026',  'balance_col': 27},
    {'name': 'февраль 2026', 'balance_col': 28},
    {'name': 'март 2026',    'balance_col': 28},
]

CLIENT_COL = 2
J_COL = 10
DATA_START = 4

# Типы которые считаются долговыми
DEBT_TYPES = {'к', 'н/к', 'п/к', 'ф'}

def normalize(name):
    if not name:
        return None
    s = str(name).strip().lower().replace('ё', 'е')
    s = ' '.join(s.split())
    return s if s else None

# ============================================================
# Сбор данных: для каждого клиента берём последний лист
# ============================================================

client_data = {}

for sheet_idx, cfg in enumerate(sheets_config):
    ws = wb[cfg['name']]
    bal_col = cfg['balance_col']

    for row in range(DATA_START, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=CLIENT_COL).value
        norm = normalize(raw_name)
        if not norm:
            continue

        j_val = ws.cell(row=row, column=J_COL).value
        j_str = str(j_val).strip().lower() if j_val is not None else ''

        # Фильтр: только долговые типы (к, н/к, п/к, ф)
        if j_str not in DEBT_TYPES:
            continue

        bal_val = ws.cell(row=row, column=bal_col).value
        if bal_val is None:
            continue
        try:
            balance = float(bal_val)
        except (ValueError, TypeError):
            continue

        row_info = {'balance': balance, 'row': row, 'j_type': j_str}

        if norm not in client_data or client_data[norm]['sheet_idx'] < sheet_idx:
            client_data[norm] = {
                'sheet_idx': sheet_idx,
                'sheet_name': cfg['name'],
                'raw_name': str(raw_name).strip(),
                'rows': [row_info],
            }
        elif client_data[norm]['sheet_idx'] == sheet_idx:
            client_data[norm]['rows'].append(row_info)

# ============================================================
# Суммируем долг по каждому клиенту
# ============================================================

results = []
for norm, data in client_data.items():
    total_debt = sum(r['balance'] for r in data['rows'])
    row_count = len(data['rows'])

    # Разбивка по типам J
    j_breakdown = {}
    for r in data['rows']:
        jt = r['j_type']
        if jt not in j_breakdown:
            j_breakdown[jt] = {'sum': 0, 'count': 0}
        j_breakdown[jt]['sum'] += r['balance']
        j_breakdown[jt]['count'] += 1

    results.append({
        'name': data['raw_name'],
        'norm': norm,
        'debt': total_debt,
        'rows': row_count,
        'sheet': data['sheet_name'],
        'j_breakdown': j_breakdown,
    })

# Фильтруем и сортируем
debtors = [r for r in results if r['debt'] > 0]
debtors.sort(key=lambda x: x['debt'], reverse=True)
non_debtors_neg = [r for r in results if r['debt'] < 0]
non_debtors_neg.sort(key=lambda x: x['debt'])

def fmt(n):
    if n == int(n):
        return f"{int(n):,}".replace(',', ' ')
    return f"{n:,.2f}".replace(',', ' ')

# ============================================================
# Общая аналитика по типам J
# ============================================================

total_by_j = {}
for d in debtors:
    for jt, info in d['j_breakdown'].items():
        if jt not in total_by_j:
            total_by_j[jt] = {'sum': 0, 'count': 0, 'clients': 0}
        total_by_j[jt]['sum'] += max(0, info['sum'])
        total_by_j[jt]['count'] += info['count']
        if info['sum'] > 0:
            total_by_j[jt]['clients'] += 1

total_debt = sum(d['debt'] for d in debtors)
total_rows = sum(d['rows'] for d in debtors)
num_debtors = len(debtors)
avg_debt = total_debt / num_debtors if num_debtors > 0 else 0

# ============================================================
# Вывод
# ============================================================

print("=" * 95)
print("  ФИНАНСОВЫЙ АУДИТ ДОЛГОВ (с учётом колонки J)")
print(f"  Источник: 03.03.2026.xlsx")
print(f"  Типы долга: к, н/к, п/к, ф")
print("=" * 95)

print(f"\n{'─' * 95}")
print(f"  СВОДКА")
print(f"{'─' * 95}")
print(f"  Общая сумма долга:          {fmt(total_debt)} сум")
print(f"  Количество должников:       {num_debtors}")
print(f"  Количество строк с долгом:  {total_rows}")
print(f"  Средний долг клиента:       {fmt(avg_debt)} сум")
if debtors:
    print(f"  Максимальный долг:          {fmt(debtors[0]['debt'])} сум ({debtors[0]['name']})")
    print(f"  Минимальный долг:           {fmt(debtors[-1]['debt'])} сум ({debtors[-1]['name']})")

print(f"\n{'─' * 95}")
print(f"  РАЗБИВКА ПО ТИПАМ ОПЛАТЫ (колонка J)")
print(f"{'─' * 95}")
j_labels = {
    'к': 'к    — заказ в долг',
    'н/к': 'н/к  — наличная оплата, заказ в долг',
    'п/к': 'п/к  — перечисление, заказ в долг',
    'ф': 'ф    — Фотих (свой человек)',
}
print(f"  {'Тип':<42} {'Сумма долга':>20} {'Строк':>7} {'Клиен.':>7}")
print(f"  {'─'*80}")
for jt in ['к', 'н/к', 'п/к', 'ф']:
    if jt in total_by_j:
        info = total_by_j[jt]
        label = j_labels.get(jt, jt)
        print(f"  {label:<42} {fmt(info['sum']):>20} {info['count']:>7} {info['clients']:>7}")

# ============================================================
# ТОП-20
# ============================================================

print(f"\n{'─' * 95}")
print(f"  ТОП-20 КРУПНЕЙШИХ ДОЛЖНИКОВ")
print(f"{'─' * 95}")
print(f"{'#':>4}  {'Клиент':<35} {'Долг (сум)':>20} {'% общ.':>7} {'Строк':>6}  {'Типы J'}")
print("─" * 95)

for i, d in enumerate(debtors[:20], 1):
    pct = (d['debt'] / total_debt * 100) if total_debt > 0 else 0
    name = d['name'][:33] + '..' if len(d['name']) > 35 else d['name']
    j_types = ', '.join(sorted(d['j_breakdown'].keys()))
    print(f"{i:>4}  {name:<35} {fmt(d['debt']):>20} {pct:>6.1f}% {d['rows']:>6}  {j_types}")

top20_total = sum(d['debt'] for d in debtors[:20])
top20_pct = (top20_total / total_debt * 100) if total_debt > 0 else 0
print("─" * 95)
print(f"{'':>4}  {'Итого ТОП-20':<35} {fmt(top20_total):>20} {top20_pct:>6.1f}%")

# ============================================================
# Полный список
# ============================================================

print(f"\n{'─' * 95}")
print(f"  ПОЛНЫЙ СПИСОК ДОЛЖНИКОВ ({num_debtors} клиентов)")
print(f"{'─' * 95}")
print(f"{'#':>4}  {'Клиент':<35} {'Долг (сум)':>20} {'Строк':>6}  {'Типы J':<15} {'Лист'}")
print("─" * 95)

for i, d in enumerate(debtors, 1):
    name = d['name'][:33] + '..' if len(d['name']) > 35 else d['name']
    j_types = ', '.join(sorted(d['j_breakdown'].keys()))
    print(f"{i:>4}  {name:<35} {fmt(d['debt']):>20} {d['rows']:>6}  {j_types:<15} {d['sheet']}")

print("─" * 95)
print(f"{'':>4}  {'ИТОГО':<35} {fmt(total_debt):>20} {total_rows:>6}")
print("=" * 95)

# ===== Отдельно: клиенты только с типом "ф" =====
f_clients = [d for d in debtors if 'ф' in d['j_breakdown']]
if f_clients:
    print(f"\n{'─' * 95}")
    print(f"  ДОЛЖНИКИ С ТИПОМ 'Ф' (Фотих) — {len(f_clients)} клиентов")
    print(f"{'─' * 95}")
    print(f"{'#':>4}  {'Клиент':<35} {'Долг (общий)':>20} {'Долг тип Ф':>20} {'Строк Ф':>8}")
    print("─" * 95)
    f_total = 0
    f_total_f_only = 0
    for i, d in enumerate(f_clients, 1):
        fi = d['j_breakdown']['ф']
        f_total += d['debt']
        f_total_f_only += fi['sum']
        name = d['name'][:33] + '..' if len(d['name']) > 35 else d['name']
        print(f"{i:>4}  {name:<35} {fmt(d['debt']):>20} {fmt(fi['sum']):>20} {fi['count']:>8}")
    print("─" * 95)
    print(f"{'':>4}  {'ИТОГО':<35} {fmt(f_total):>20} {fmt(f_total_f_only):>20}")

# ===== Переплаты =====
if non_debtors_neg:
    print(f"\n{'─' * 95}")
    print(f"  КЛИЕНТЫ С ПЕРЕПЛАТОЙ ({len(non_debtors_neg)})")
    print(f"{'─' * 95}")
    for d in non_debtors_neg:
        j_types = ', '.join(sorted(d['j_breakdown'].keys()))
        print(f"  {d['name']:<35} {fmt(d['debt']):>20} сум  ({j_types})")
    neg_total = sum(d['debt'] for d in non_debtors_neg)
    print(f"  {'ИТОГО переплата':<35} {fmt(neg_total):>20} сум")

print(f"\n{'=' * 95}")
print(f"  ФОРМУЛА")
print(f"{'=' * 95}")
print(f"  Фильтр строк:  колонка J IN ('к', 'н/к', 'п/к', 'ф')")
print(f"  Долг строки:    колонка остатка (AA/27 янв, AB/28 фев/мар)")
print(f"  Debt(клиент) =  SUM(остаток) всех строк клиента в последнем листе")
print(f"  TotalDebt    =  SUM(Debt) для всех клиентов где Debt > 0")
print("=" * 95)
