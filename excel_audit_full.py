"""
ПОЛНЫЙ EXCEL-АУДИТ CRM
=======================
Версия: 2.0  |  Дата: 2026-03-09
Статус: production-ready dry-run

Изменения v2.0 (09.03.2026):
  - Фиксированная карта маппинга колонок по (год, месяц)
  - Расширенные credit_codes: к, н/к, п/к, пк/к, ф, фт, фотих
  - Предоплата (ПП): детекция prepay_sum, вычисление net_debt
  - Новые CSV-колонки: remaining, prepay_sum, net_debt, taken_on_credit_flag
  - Per-client агрегация по net_debt для reconciliation

Этапы:
  0) Бэкап уже сделан (raw_backup/). Baseline в mnt/data/baseline_debt.json
  1) Полное сканирование каждого файла/листа
  2) Определение колонок: фикс-мэп по (год,месяц) + fallback автодетекция
  3) Нормализация: orders_clean, payments_normalized
  4) Аномалии: пустые даты, смешанные форматы, отрицательные остатки, дубли
  5) Агрегация по клиентам: remaining, prepay_sum, net_debt, taken_on_credit
  6) Реконсилиэйшн Excel ↔ CRM
  7) Экспорт CSV + mapping JSON + Excel-отчёт
  8) Верификация: baseline_debt не изменился (только чтение!)
"""

import os, sys, json, re, csv, logging
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import openpyxl
import psycopg2

# ──── НАСТРОЙКИ ────────────────────────────────────────────────────────────────
ROOT = Path(r"c:\Users\Noutbuk savdosi\CRM")
OUT  = ROOT / "mnt"

FILES = {
    "03.03.2026": ROOT / "03.03.2026.xlsx",
    "29.12.2025": ROOT / "29.12.2025.xlsx",
    "26.12.2024": ROOT / "26.12.2024.xlsx",
}

DB_URL = (
    "postgresql://crm_user:BChpe9Gb4dOeVQQxRYVkiLUgu4TsmWJo"
    "@dpg-d6bcdrt6ubrc73ch10dg-a.oregon-postgres.render.com/crm_db_okj8"
)

# ──── КОДЫ ТИПОВ ОПЛАТЫ ──────────────────────────────────────────────────────

# Кредитные коды (= долг)
CREDIT_CODES = {"к", "н/к", "п/к", "пк/к", "ф", "фт", "фотих"}
# Предоплатные коды
PREPAY_CODES = {"пп"}
# Все известные J-коды (для автодетекции колонки J)
ALL_J_CODES  = CREDIT_CODES | PREPAY_CODES | {"н", "п", "нал", "нкл", "б/н"}
ALL_J_CODES_LOWER = {x.lower() for x in ALL_J_CODES}

# Для обратной совместимости
DEBT_J_CODES = CREDIT_CODES

TS = datetime.now().strftime("%Y%m%d_%H%M%S")
TOLERANCE = 1_000  # расхождение < 1000 сум считается OK

# ──── ФИКСИРОВАННАЯ КАРТА МАППИНГА КОЛОНОК ──────────────────────────────────

# Ключ: (year_str, month_lower) → dict с j_col (int) и balance_col (int)
# j_col и balance_col — 1-based номера колонок Excel
# J=10 (col J), L=12, M=13, AA=27, AB=28, AC=29, AF=32

MONTH_NAMES_RU = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}

def _build_fixed_mapping():
    """Строит полную таблицу маппинга (year, month_num) → {j_col, balance_col}."""
    m = {}

    # 2026
    m[("2026", 1)]  = {"j_col": 10, "balance_col": 27}  # Январь: J, AA
    m[("2026", 2)]  = {"j_col": 10, "balance_col": 28}  # Февраль: J, AB
    m[("2026", 3)]  = {"j_col": 10, "balance_col": 28}  # Март: J, AB

    # 2025: все месяцы J=10, остаток AA=27
    for mo in range(1, 13):
        m[("2025", mo)] = {"j_col": 10, "balance_col": 27}

    # 2024
    # Январь: M(13)=payment_type, остаток AF(32)
    m[("2024", 1)]  = {"j_col": 13, "balance_col": 32}
    # Февраль → Октябрь: L(12)=payment_type, остаток AC(29)
    for mo in range(2, 11):
        m[("2024", mo)] = {"j_col": 12, "balance_col": 29}
    # Ноябрь, Декабрь: J(10)=payment_type, остаток AA(27)
    m[("2024", 11)] = {"j_col": 10, "balance_col": 27}
    m[("2024", 12)] = {"j_col": 10, "balance_col": 27}

    return m

FIXED_MAPPING = _build_fixed_mapping()


def _parse_year_from_filekey(file_key):
    """Извлекает год из file_key вида '03.03.2026'."""
    parts = file_key.split(".")
    if len(parts) == 3 and parts[2].isdigit():
        return parts[2]
    return None


def _parse_month_from_sheetname(sheet_name):
    """Извлекает номер месяца из имени листа (на русском)."""
    sn = sheet_name.strip().lower()
    for name, num in MONTH_NAMES_RU.items():
        if name in sn:
            return num
    return None


def get_column_config(file_key, sheet_name, ws):
    """
    Определяет j_col и balance_col для листа.
    1) Пробуем фикс-мэп по (год, месяц)
    2) Если не найдено — fallback на автодетекцию
    """
    year = _parse_year_from_filekey(file_key)
    month = _parse_month_from_sheetname(sheet_name)

    cfg_source = "autodetect"
    if year and month and (year, month) in FIXED_MAPPING:
        fm = FIXED_MAPPING[(year, month)]
        cfg_source = "fixed_map"
        # data_start всё ещё определяем автоматически
        data_start = _detect_data_start(ws)
        return {
            "client_col":  2,
            "j_col":       fm["j_col"],
            "balance_col": fm["balance_col"],
            "data_start":  data_start,
            "_balance_hdr": f"fixed_map(year={year},month={month})",
            "_cfg_source": cfg_source,
        }

    # Fallback: автодетекция
    cfg = autodetect_cfg(ws)
    cfg["_cfg_source"] = cfg_source
    return cfg


# ──── ЛОГИРОВАНИЕ ──────────────────────────────────────────────────────────────
log_dir = OUT / "logs"
log_dir.mkdir(parents=True, exist_ok=True)
log_path = log_dir / f"cleaning_run_{TS}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(log_path, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("audit")

# ──── УТИЛИТЫ ──────────────────────────────────────────────────────────────────

def normalize_name(s):
    """Нормализация: lower, strip, ё→е, collapse spaces."""
    if s is None:
        return None
    s = str(s).strip().lower()
    s = s.replace("ё", "е")
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip()
    return s if s else None


def parse_number(val):
    """Привести ячейку к числу. Вернуть (float|None, anomaly_flag)."""
    if val is None:
        return None, False
    if isinstance(val, (int, float)):
        return float(val), False
    s = str(val).strip()
    s = re.sub(r"[\s\u00a0\u202f]", "", s)
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s:
        return None, True
    try:
        return float(s), False
    except ValueError:
        return None, True


def levenshtein(a, b):
    """Простое расстояние Левенштейна."""
    if len(a) < len(b):
        a, b = b, a
    if len(b) == 0:
        return len(a)
    row = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        new_row = [i + 1]
        for j, cb in enumerate(b):
            new_row.append(min(row[j + 1] + 1, new_row[j] + 1, row[j] + (ca != cb)))
        row = new_row
    return row[-1]


def fuzzy_match_clients(names):
    """Возвращает список предложений по слиянию (fuzzy). Не объединяет автоматически."""
    norm_list = list(set(n for n in names if n))
    suggestions = []
    for i in range(len(norm_list)):
        for j in range(i + 1, len(norm_list)):
            a, b = norm_list[i], norm_list[j]
            max_len = max(len(a), len(b), 1)
            dist = levenshtein(a, b)
            threshold = 2 if max_len <= 10 else max(2, int(max_len * 0.10))
            if dist <= threshold:
                suggestions.append((a, b, dist))
            # Word-sort check: "носир кредо" ↔ "кредо носир"
            elif " ".join(sorted(a.split())) == " ".join(sorted(b.split())):
                suggestions.append((a, b, 0))  # distance=0 for word-sorted match
    return suggestions


# ──── АВТОДЕТЕКЦИЯ КОЛОНОК (FALLBACK) ─────────────────────────────────────────

def _detect_data_start(ws):
    """Находит первую строку данных (не заголовок) по столбцу B."""
    max_r = ws.max_row or 10
    for row in range(1, min(10, max_r + 1)):
        v = ws.cell(row=row, column=2).value
        if v and len(str(v).strip()) > 1 and not any(
            kw in str(v).lower() for kw in ["фирм", "клиент", "назв", "дата", "ост", "номер", "#"]
        ):
            return row
    return 4  # default


def autodetect_cfg(ws):
    """
    Автодетекция j_col, balance_col, data_start для любого листа.
    Используется как fallback, если фикс-мэп не найден.
    """
    max_c = ws.max_column or 30
    max_r = ws.max_row or 10

    data_start = _detect_data_start(ws)

    # j_col: колонка где наибольшее кол-во значений из ALL_J_CODES_LOWER
    j_votes = {}
    for row in range(data_start, min(data_start + 40, max_r + 1)):
        for col in range(1, max_c + 1):
            v = ws.cell(row=row, column=col).value
            if v and str(v).strip().lower() in ALL_J_CODES_LOWER:
                j_votes[col] = j_votes.get(col, 0) + 1
    j_col = max(j_votes, key=j_votes.get) if j_votes else 10

    # balance_col: последняя колонка где заголовок содержит 'ост' и col > 15
    balance_col = None
    balance_col_header = ""
    for row in range(1, min(4, max_r + 1)):
        for col in range(16, max_c + 1):
            v = ws.cell(row=row, column=col).value
            if v and any(k in str(v).lower() for k in ["ост", "остат", "ост-к", "ост_на"]):
                if balance_col is None or col > balance_col:
                    balance_col = col
                    balance_col_header = str(v)
    if balance_col is None:
        balance_col = max_c

    return {
        "client_col":  2,
        "j_col":       j_col,
        "balance_col": balance_col,
        "data_start":  data_start,
        "_balance_hdr": balance_col_header,
    }


# Листы-пустышки
SKIP_SHEETS = {"Лист1", "Лист2", "лист1", "лист2"}


# ──── ЭТАП 1: ПОЛНЫЙ СКАН ФАЙЛОВ ───────────────────────────────────────────────

def scan_sheet(wb_name, ws, cfg):
    """
    Сканирует лист: возвращает dict с raw_rows, column_roles, anomalies.
    cfg = {client_col, balance_col, data_start, j_col}
    """
    client_col  = cfg.get("client_col", 2)
    j_col       = cfg.get("j_col", 10)
    balance_col = cfg["balance_col"]
    data_start  = cfg.get("data_start", 4)

    # Получим заголовки
    header_row = max(1, data_start - 1)
    headers = {}
    for col in range(1, (ws.max_column or 30) + 1):
        v = ws.cell(row=header_row, column=col).value
        headers[col] = str(v).strip() if v is not None else f"col_{col}"

    # Автообнаружение ролей колонок
    col_roles = {}
    col_roles[client_col]  = {"role": "client",        "name": headers.get(client_col, "B"),  "confidence": "high"}
    col_roles[j_col]       = {"role": "payment_type",  "name": headers.get(j_col, "J"),       "confidence": "high"}
    col_roles[balance_col] = {"role": "remaining_debt", "name": headers.get(balance_col, "?"), "confidence": "high"}

    # Дата
    for col, hdr in headers.items():
        hl = hdr.lower()
        if any(k in hl for k in ["дата", "date", "число"]):
            col_roles[col] = {"role": "order_date", "name": hdr, "confidence": "high"}
            break

    # Менеджер
    for col, hdr in headers.items():
        if any(k in hdr.lower() for k in ["менедж", "manager", "менеджер"]):
            col_roles[col] = {"role": "manager", "name": hdr, "confidence": "high"}
            break

    # Платёжные слоты (каждые 3 колонки между j_col+1 и balance_col-1)
    payment_slots = []
    slot_start = j_col + 1
    slot_end   = balance_col - 1
    if slot_start <= slot_end:
        slot_cols = list(range(slot_start, slot_end + 1))
        for i in range(0, len(slot_cols) - 1, 3):
            chunk = slot_cols[i:i+3]
            if len(chunk) >= 2:
                payment_slots.append({
                    "index": i // 3,
                    "sum_col":    chunk[0],
                    "method_col": chunk[1] if len(chunk) > 1 else None,
                    "note_col":   chunk[2] if len(chunk) > 2 else None,
                })

    cfg_src = cfg.get("_cfg_source", "?")
    log.info(f"  [{wb_name} / {ws.title}] cfg_src={cfg_src}, j_col={j_col}, "
             f"balance_col={balance_col}, {len(payment_slots)} payment slots "
             f"(cols {slot_start}-{slot_end}), data_start={data_start}")

    # Читаем строки
    raw_rows = []
    anomalies = []

    for row in range(data_start, (ws.max_row or data_start) + 1):
        raw_name = ws.cell(row=row, column=client_col).value
        norm_nm  = normalize_name(raw_name)
        if not norm_nm:
            continue

        j_raw = ws.cell(row=row, column=j_col).value
        j_str = str(j_raw).strip().lower() if j_raw is not None else ""

        bal_raw = ws.cell(row=row, column=balance_col).value
        balance, bal_anom = parse_number(bal_raw)

        if bal_anom:
            anomalies.append({
                "wb": wb_name, "sheet": ws.title, "row": row,
                "type": "non_numeric_balance",
                "col": balance_col, "value": str(bal_raw),
                "client": str(raw_name)
            })

        if balance is not None and balance < 0:
            anomalies.append({
                "wb": wb_name, "sheet": ws.title, "row": row,
                "type": "negative_balance",
                "col": balance_col, "value": balance,
                "client": str(raw_name)
            })

        # Дата заказа
        date_col = next((c for c, r in col_roles.items() if r["role"] == "order_date"), 1)
        date_raw = ws.cell(row=row, column=date_col).value
        if date_raw is None and balance is not None and balance != 0:
            anomalies.append({
                "wb": wb_name, "sheet": ws.title, "row": row,
                "type": "missing_date_with_balance",
                "value": balance, "client": str(raw_name)
            })

        # Парсим платёжные слоты
        payments = []
        for slot in payment_slots:
            sum_raw  = ws.cell(row=row, column=slot["sum_col"]).value
            meth_raw = ws.cell(row=row, column=slot["method_col"]).value if slot["method_col"] else None
            note_raw = ws.cell(row=row, column=slot["note_col"]).value   if slot["note_col"]   else None

            psum, panom = parse_number(sum_raw)

            if psum is None and sum_raw is not None:
                s = str(sum_raw).strip().lower() if sum_raw else ""
                if s in ALL_J_CODES_LOWER:
                    anomalies.append({
                        "wb": wb_name, "sheet": ws.title, "row": row,
                        "type": "ambiguous_payment_cell",
                        "col": slot["sum_col"], "value": str(sum_raw),
                        "client": str(raw_name), "slot_index": slot["index"]
                    })
                    continue

            if psum is not None and psum != 0:
                method_str = str(meth_raw).strip().lower() if meth_raw else ""
                payments.append({
                    "slot_index": slot["index"],
                    "amount":     psum,
                    "method":     str(meth_raw).strip() if meth_raw else None,
                    "method_lower": method_str,
                    "note":       str(note_raw).strip() if note_raw else None,
                    "raw_sum":    sum_raw, "raw_method": meth_raw, "raw_note": note_raw,
                    "sum_col":    slot["sum_col"],
                })

        # Остальные поля
        manager_col = next((c for c, r in col_roles.items() if r["role"] == "manager"), None)
        manager_raw = ws.cell(row=row, column=manager_col).value if manager_col else None

        raw_rows.append({
            "file":        wb_name,
            "sheet":       ws.title,
            "row_id":      row,
            "client_raw":  str(raw_name).strip() if raw_name else "",
            "client_norm": norm_nm,
            "j_raw":       j_raw,
            "j_str":       j_str,
            "balance_raw": bal_raw,
            "balance":     balance,
            "date_raw":    date_raw,
            "manager_raw": manager_raw,
            "payments":    payments,
        })

    log.info(f"  [{wb_name} / {ws.title}] rows={len(raw_rows)}, anomalies={len(anomalies)}")
    return {
        "wb_name":       wb_name,
        "sheet":         ws.title,
        "headers":       headers,
        "col_roles":     col_roles,
        "payment_slots": payment_slots,
        "raw_rows":      raw_rows,
        "anomalies":     anomalies,
    }


def scan_all_files():
    all_scans = []
    for file_key, fpath in FILES.items():
        if not fpath.exists():
            log.error(f"FILE NOT FOUND: {fpath}")
            continue
        log.info(f"\n{'='*60}")
        log.info(f"SCANNING: {file_key} ({fpath.name})")
        log.info(f"{'='*60}")
        wb = openpyxl.load_workbook(str(fpath), data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                log.info(f"  SKIP (служебный лист): '{sheet_name}'")
                continue
            log.info(f"\n  Sheet: '{sheet_name}'")
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row < 5:
                log.info(f"  SKIP (пустой лист, max_row={ws.max_row})")
                continue
            # Определяем колонки: фикс-мэп + fallback
            cfg = get_column_config(file_key, sheet_name, ws)
            cfg_src = cfg.get("_cfg_source", "?")
            log.info(f"    Column cfg ({cfg_src}): j_col={cfg['j_col']}, "
                     f"balance_col={cfg['balance_col']} "
                     f"('{cfg.get('_balance_hdr','')}'), "
                     f"data_start={cfg['data_start']}")
            try:
                scan = scan_sheet(file_key, ws, cfg)
                scan["file_key"] = file_key
                all_scans.append(scan)
            except Exception as e:
                log.error(f"  ERROR scanning {file_key}/{sheet_name}: {e}")
                import traceback; log.error(traceback.format_exc())
        wb.close()
    return all_scans


# ──── ЭТАП 2: НОРМАЛИЗАЦИЯ ORDERS + PAYMENTS ───────────────────────────────────

def _detect_prepay_in_payments(payments_list):
    """
    Суммирует предоплатные платежи из payment slots.
    Если метод содержит ПП, это предоплата.
    Возвращает prepay_sum.
    """
    prepay_sum = 0.0
    for pmt in payments_list:
        method = pmt.get("method_lower", "")
        note   = (pmt.get("note") or "").lower()
        if method in PREPAY_CODES or "пп" in method or "пп" in note:
            prepay_sum += abs(pmt["amount"])
    return prepay_sum


def build_orders_and_payments(all_scans):
    """
    Логика: для каждого клиента берём ПОСЛЕДНИЙ файл + ПОСЛЕДНИЙ лист
    где он присутствует. Строки с J IN CREDIT_CODES + prepay detection.
    """
    FILE_ORDER = ["26.12.2024", "29.12.2025", "03.03.2026"]

    client_appearances = defaultdict(list)
    scan_order = {}
    for fi, fk in enumerate(FILE_ORDER):
        sheet_idx = 0
        for scan in all_scans:
            if scan["file_key"] == fk:
                scan_order[(fk, scan["sheet"])] = (fi, sheet_idx)
                sheet_idx += 1
                for row in scan["raw_rows"]:
                    client_appearances[row["client_norm"]].append(
                        (fi, sheet_idx - 1, scan, row)
                    )

    orders = []
    payments_normalized = []
    order_id = 0

    for client_norm, appearances in client_appearances.items():
        if not appearances:
            continue
        last_fi, last_si, last_scan, _ = max(appearances, key=lambda x: (x[0], x[1]))
        rows_in_last = [r for fi, si, sc, r in appearances if fi == last_fi and si == last_si]

        for row in rows_in_last:
            order_id += 1
            j_lower = row["j_str"]
            is_credit = j_lower in CREDIT_CODES
            is_prepay_type = j_lower in PREPAY_CODES

            # Предоплата из платёжных слотов
            prepay_from_slots = _detect_prepay_in_payments(row["payments"])

            # Если сама строка — ПП, то remaining_debt может быть предоплатой
            prepay_sum = prepay_from_slots
            if is_prepay_type:
                # Сумма ПП-строки = её balance (если > 0, это prepayment)
                bal = row["balance"] if row["balance"] is not None else 0.0
                if bal > 0:
                    prepay_sum += bal

            remaining = row["balance"] if row["balance"] is not None else 0.0
            # net_debt рассчитывается потом на уровне per-client, не per-row,
            # но для CSV выводим per-row remaining
            prepayment_flag = 1 if (is_prepay_type or prepay_from_slots > 0) else 0

            orders.append({
                "order_id":        order_id,
                "file":            row["file"],
                "sheet":           row["sheet"],
                "row_id":          row["row_id"],
                "client_raw":      row["client_raw"],
                "client_norm":     row["client_norm"],
                "manager_raw":     row["manager_raw"],
                "order_date_raw":  str(row["date_raw"]) if row["date_raw"] else "",
                "j_raw":           str(row["j_raw"]) if row["j_raw"] else "",
                "j_str":           j_lower,
                "is_debt_type":    is_credit,
                "is_prepay_type":  is_prepay_type,
                "balance_raw":     str(row["balance_raw"]) if row["balance_raw"] is not None else "",
                "remaining_debt":  remaining,
                "prepay_sum":      prepay_sum,
                "prepayment_flag": prepayment_flag,
                "taken_on_credit_flag": 1 if is_credit else 0,
                "payments_count":  len(row["payments"]),
                "payments_raw_json": json.dumps(
                    [{"idx": p["slot_index"], "amt": p["amount"],
                      "method": p.get("method"), "note": p.get("note")}
                     for p in row["payments"]],
                    ensure_ascii=False
                ),
            })

            for pmt in row["payments"]:
                payments_normalized.append({
                    "order_id":       order_id,
                    "file":           row["file"],
                    "sheet":          row["sheet"],
                    "row_id":         row["row_id"],
                    "client_raw":     row["client_raw"],
                    "client_norm":    row["client_norm"],
                    "slot_index":     pmt["slot_index"],
                    "payment_amount": pmt["amount"],
                    "payment_method": pmt["method"],
                    "payment_note":   pmt["note"],
                    "is_prepay":      1 if (pmt.get("method_lower", "") in PREPAY_CODES
                                            or "пп" in pmt.get("method_lower", "")) else 0,
                })

    return orders, payments_normalized


# ──── ЭТАП 3: PER-CLIENT АГРЕГАЦИЯ (с NetDebt) ────────────────────────────────

def aggregate_excel_debts(orders):
    """
    Для каждого клиента:
    - remaining: SUM(remaining_debt) для credit-type строк (remaining > 0)
    - prepay_sum: SUM(prepay_sum) по всем строкам клиента
    - taken_on_credit_sum: SUM(remaining_debt) для строк с is_debt_type=True
    - net_debt: remaining - prepay_sum
    Возвращает dict: client_norm → {...}
    """
    per_client = defaultdict(lambda: {
        "remaining": 0.0,
        "prepay_sum": 0.0,
        "taken_on_credit_sum": 0.0,
        "row_count": 0,
    })

    for o in orders:
        cn = o["client_norm"]
        per_client[cn]["row_count"] += 1

        # Остаток: суммируем только credit-type строки с positive remaining
        if o["is_debt_type"] and o["remaining_debt"] > 0:
            per_client[cn]["remaining"] += o["remaining_debt"]

        # Предоплатные суммы: суммируем по всем строкам клиента
        per_client[cn]["prepay_sum"] += o["prepay_sum"]

        # Взято в долг (credit-type): суммируем remaining_debt если credit
        if o["is_debt_type"]:
            per_client[cn]["taken_on_credit_sum"] += max(0, o["remaining_debt"])

    result = {}
    for cn, data in per_client.items():
        net_debt = data["remaining"] - data["prepay_sum"]
        result[cn] = {
            "remaining":           data["remaining"],
            "prepay_sum":          data["prepay_sum"],
            "taken_on_credit_sum": data["taken_on_credit_sum"],
            "net_debt":            net_debt,
            "row_count":           data["row_count"],
        }
    return result


# ──── ЭТАП 4: CRM-ДАННЫЕ ─────────────────────────────────────────────────────

def load_crm_debts():
    """Загружаем долги из CRM по клиентам."""
    conn = psycopg2.connect(DB_URL, connect_timeout=30)
    cur = conn.cursor()

    cur.execute("""
        SELECT
          c.company_name, c.id,
          SUM(d.amount - d.paid_amount) AS debt,
          COUNT(d.id) AS deal_count
        FROM deals d
        JOIN clients c ON d.client_id = c.id
        WHERE d.payment_status IN ('UNPAID', 'PARTIAL')
          AND d.status NOT IN ('CANCELED', 'REJECTED')
          AND d.is_archived = false
        GROUP BY c.id, c.company_name
        HAVING SUM(d.amount - d.paid_amount) > 0
        ORDER BY debt DESC
    """)
    crm_clients = {}
    for row in cur.fetchall():
        company_name, cid, debt, dcnt = row
        norm = normalize_name(company_name)
        crm_clients[norm] = {
            "id":           str(cid),
            "company_name": company_name,
            "debt":         float(debt),
            "deal_count":   dcnt,
            "norm":         norm,
        }

    cur.execute("""
        SELECT COALESCE(SUM(d.amount - d.paid_amount), 0)
        FROM deals d
        WHERE d.payment_status IN ('UNPAID', 'PARTIAL')
          AND d.status NOT IN ('CANCELED', 'REJECTED')
          AND d.is_archived = false
          AND (d.amount - d.paid_amount) > 0
    """)
    total_crm = float(cur.fetchone()[0])

    cur.close(); conn.close()
    return crm_clients, total_crm


# ──── ЭТАП 5: РЕКОНСИЛИЭЙШН ───────────────────────────────────────────────────

def reconcile(excel_agg, crm_clients, total_crm):
    """
    Сравниваем net_debt из Excel с CRM.
    excel_agg: client_norm → {remaining, prepay_sum, net_debt, taken_on_credit_sum, ...}
    """
    diffs = []
    excel_names = set(excel_agg.keys())
    crm_names   = set(crm_clients.keys())

    both = excel_names & crm_names
    only_crm_set   = crm_names - excel_names
    only_excel_set = excel_names - crm_names

    for norm in sorted(both):
        ea = excel_agg[norm]
        excel_remaining = ea["remaining"]
        excel_net = ea["net_debt"]
        crm_debt = crm_clients[norm]["debt"]
        diff = crm_debt - excel_remaining  # используем remaining для основного сравнения
        diff_net = crm_debt - max(0, excel_net)
        if abs(diff) >= TOLERANCE:
            status = "CRM_HIGHER" if diff > 0 else "EXCEL_HIGHER"
            diffs.append({
                "client_norm":       norm,
                "company_name":      crm_clients[norm]["company_name"],
                "excel_remaining":   excel_remaining,
                "excel_prepay":      ea["prepay_sum"],
                "excel_net_debt":    max(0, excel_net),
                "excel_credit_sum":  ea["taken_on_credit_sum"],
                "crm_debt":          crm_debt,
                "diff":              diff,
                "diff_net":          diff_net,
                "status":            status,
                "root_cause":        _guess_root_cause(diff, crm_clients[norm]),
                # Backward compatibility
                "excel_debt":        excel_remaining,
            })

    diffs.sort(key=lambda x: abs(x["diff"]), reverse=True)

    only_crm = [
        {"client_norm": n, "company_name": crm_clients[n]["company_name"],
         "crm_debt": crm_clients[n]["debt"], "deal_count": crm_clients[n]["deal_count"]}
        for n in sorted(only_crm_set)
    ]
    only_excel = []
    for n in sorted(only_excel_set):
        ea = excel_agg[n]
        if ea["remaining"] > 0:
            only_excel.append({
                "client_norm": n,
                "excel_debt": ea["remaining"],
                "excel_net_debt": max(0, ea["net_debt"]),
                "excel_prepay": ea["prepay_sum"],
            })

    total_excel = sum(ea["remaining"] for ea in excel_agg.values() if ea["remaining"] > 0)
    total_excel_net = sum(max(0, ea["net_debt"]) for ea in excel_agg.values())
    total_prepay = sum(ea["prepay_sum"] for ea in excel_agg.values())
    global_diff = total_crm - total_excel

    return {
        "total_crm":       total_crm,
        "total_excel":     total_excel,
        "total_excel_net": total_excel_net,
        "total_prepay":    total_prepay,
        "global_diff":     global_diff,
        "diffs":           diffs,
        "only_crm":        only_crm,
        "only_excel":      only_excel,
    }


def _guess_root_cause(diff, crm_info):
    if diff > 50_000_000:
        return "Возможно: ненормализованное имя клиента или несколько сделок не в Excel"
    if diff > 5_000_000:
        return "Скорее всего: сделки 2024-2025 не закрыты в CRM или частично оплачены"
    if diff < -1_000_000:
        return "Excel содержит долг, которого нет в CRM (возможно, архивирована сделка)"
    return "Незначительное расхождение -- проверить вручную"


# ──── ЭТАП 6: ЭКСПОРТ ──────────────────────────────────────────────────────────

def export_csvs(all_scans, orders, payments_normalized, anomalies_all, reconciliation, excel_agg):
    cleaned_root = OUT / "data" / "cleaned"
    reports_root = OUT / "data" / "reports"
    cleaned_root.mkdir(parents=True, exist_ok=True)
    reports_root.mkdir(parents=True, exist_ok=True)

    # 1) CSV orders per file+sheet
    orders_by_fs = defaultdict(list)
    payments_by_fs = defaultdict(list)
    for o in orders:
        orders_by_fs[(o["file"], o["sheet"])].append(o)
    for p in payments_normalized:
        payments_by_fs[(p["file"], p["sheet"])].append(p)

    for (fk, sh), rows in orders_by_fs.items():
        d = cleaned_root / fk
        d.mkdir(exist_ok=True)
        safe_sh = sh.replace(" ", "_").replace("/", "-")
        path = d / f"orders_clean_{safe_sh}.csv"
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            if rows:
                w = csv.DictWriter(f, fieldnames=rows[0].keys())
                w.writeheader(); w.writerows(rows)
        log.info(f"  CSV orders   -> {path}")

    for (fk, sh), rows in payments_by_fs.items():
        d = cleaned_root / fk
        d.mkdir(exist_ok=True)
        safe_sh = sh.replace(" ", "_").replace("/", "-")
        path = d / f"payments_normalized_{safe_sh}.csv"
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            if rows:
                w = csv.DictWriter(f, fieldnames=rows[0].keys())
                w.writeheader(); w.writerows(rows)
        log.info(f"  CSV payments -> {path}")

    # 2) Mapping JSON per file+sheet
    for scan in all_scans:
        fk = scan["file_key"]
        sh = scan["sheet"]
        d = cleaned_root / fk
        d.mkdir(exist_ok=True)
        safe_sh = sh.replace(" ", "_").replace("/", "-")
        mapping = {
            str(col): {
                "original_header": scan["headers"].get(col, ""),
                "role":            info["role"],
                "confidence":      info["confidence"],
            }
            for col, info in scan["col_roles"].items()
        }
        path = d / f"mapping_{safe_sh}.json"
        with open(path, "w", encoding="utf-8") as f:
            json.dump(mapping, f, ensure_ascii=False, indent=2)
        log.info(f"  JSON mapping -> {path}")

    # 3) Anomalies CSV (collect all possible field names across all anomaly types)
    anom_path = reports_root / f"anomalies_{TS}.csv"
    with open(anom_path, "w", newline="", encoding="utf-8-sig") as f:
        if anomalies_all:
            all_keys = []
            seen = set()
            for a in anomalies_all:
                for k in a.keys():
                    if k not in seen:
                        all_keys.append(k)
                        seen.add(k)
            w = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
            w.writeheader(); w.writerows(anomalies_all)
    log.info(f"  Anomalies CSV -> {anom_path} ({len(anomalies_all)} records)")

    # 4) Per-client aggregation CSV
    agg_path = reports_root / f"per_client_agg_{TS}.csv"
    with open(agg_path, "w", newline="", encoding="utf-8-sig") as f:
        fields = ["client_norm", "remaining", "prepay_sum", "net_debt",
                  "taken_on_credit_sum", "row_count", "display"]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for cn in sorted(excel_agg.keys()):
            ea = excel_agg[cn]
            nd = ea["net_debt"]
            if nd > 0:
                display = f"Долг: {nd:,.0f}"
            elif nd < 0:
                display = f"Переплата: {abs(nd):,.0f}"
            else:
                display = "0"
            w.writerow({
                "client_norm":       cn,
                "remaining":         ea["remaining"],
                "prepay_sum":        ea["prepay_sum"],
                "net_debt":          ea["net_debt"],
                "taken_on_credit_sum": ea["taken_on_credit_sum"],
                "row_count":         ea["row_count"],
                "display":           display,
            })
    log.info(f"  Per-client agg -> {agg_path}")

    # 5) Reconciliation XLSX
    export_reconciliation_xlsx(reconciliation, anomalies_all, reports_root)


def export_reconciliation_xlsx(rec, anomalies_all, reports_root):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    RED   = PatternFill("solid", fgColor="FFCCCC")
    YEL   = PatternFill("solid", fgColor="FFFACD")
    GRN   = PatternFill("solid", fgColor="CCFFCC")
    BLUE  = PatternFill("solid", fgColor="CCE0FF")
    ORG   = PatternFill("solid", fgColor="FFE0B2")
    HDR   = PatternFill("solid", fgColor="2F5496")
    HFONT = Font(color="FFFFFF", bold=True)

    def header_row(ws, cols):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.fill = HDR; cell.font = HFONT; cell.alignment = Alignment(horizontal="center")

    def autowidth(ws, extra=4):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + extra, 60)

    # ── Sheet 1: Summary
    ws = wb.active
    ws.title = "summary"
    ws.append(["АУДИТ EXCEL vs CRM v2.0", f"Сформировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["Показатель", "Значение"])
    summary_data = [
        ("Итого долг в CRM (baseline)",         f"{rec['total_crm']:,.2f} сум"),
        ("Итого остаток Excel (remaining)",      f"{rec['total_excel']:,.2f} сум"),
        ("Итого предоплат Excel (prepay_sum)",   f"{rec['total_prepay']:,.2f} сум"),
        ("Итого net_debt Excel",                 f"{rec['total_excel_net']:,.2f} сум"),
        ("Разница (CRM - Excel remaining)",      f"{rec['global_diff']:,.2f} сум"),
        ("Клиентов только в CRM",                len(rec["only_crm"])),
        ("Клиентов только в Excel",              len(rec["only_excel"])),
        ("Клиентов с расхождением > 1000 сум",   len(rec["diffs"])),
        ("Всего аномалий",                       len(anomalies_all)),
        ("Tolerance (порог)",                    f"{TOLERANCE:,} сум"),
    ]
    for row_data in summary_data:
        ws.append(list(row_data))
    autowidth(ws)

    # ── Sheet 2: diffs (с net_debt)
    ws2 = wb.create_sheet("diffs")
    header_row(ws2, ["#", "Клиент (norm)", "Имя в CRM",
                     "Excel Remaining", "Excel Prepay", "Excel NetDebt",
                     "CRM Долг", "Разница (CRM-Excel)", "Статус", "Причина"])
    for i, d in enumerate(rec["diffs"], 1):
        ws2.append([
            i, d["client_norm"], d["company_name"],
            round(d["excel_remaining"], 2), round(d["excel_prepay"], 2),
            round(d["excel_net_debt"], 2),
            round(d["crm_debt"], 2), round(d["diff"], 2),
            d["status"], d["root_cause"],
        ])
        fill = RED if d["status"] == "CRM_HIGHER" else YEL
        for cell in ws2[ws2.max_row]:
            cell.fill = fill
    autowidth(ws2)

    # ── Sheet 3: top20_diffs
    ws3 = wb.create_sheet("top20_diffs")
    header_row(ws3, ["#", "Клиент", "Excel Remaining", "Excel NetDebt",
                     "CRM Долг", "Разница", "Статус"])
    for i, d in enumerate(rec["diffs"][:20], 1):
        ws3.append([
            i, d.get("company_name", d["client_norm"]),
            round(d["excel_remaining"], 2), round(d["excel_net_debt"], 2),
            round(d["crm_debt"], 2), round(d["diff"], 2), d["status"],
        ])
        cell_fill = RED if d["diff"] > 0 else YEL
        for cell in ws3[ws3.max_row]:
            cell.fill = cell_fill
    autowidth(ws3)

    # ── Sheet 4: clients_only_in_CRM
    ws4 = wb.create_sheet("clients_only_in_CRM")
    header_row(ws4, ["#", "Клиент (norm)", "Имя в CRM", "Долг CRM", "Сделок"])
    for i, c in enumerate(rec["only_crm"], 1):
        ws4.append([i, c["client_norm"], c["company_name"],
                    round(c["crm_debt"], 2), c["deal_count"]])
        for cell in ws4[ws4.max_row]:
            cell.fill = BLUE
    autowidth(ws4)

    # ── Sheet 5: clients_only_in_Excel
    ws5 = wb.create_sheet("clients_only_in_Excel")
    header_row(ws5, ["#", "Клиент (norm)", "Долг Excel", "NetDebt Excel", "Prepay"])
    for i, c in enumerate(rec["only_excel"], 1):
        ws5.append([i, c["client_norm"], round(c["excel_debt"], 2),
                    round(c.get("excel_net_debt", 0), 2),
                    round(c.get("excel_prepay", 0), 2)])
        for cell in ws5[ws5.max_row]:
            cell.fill = GRN
    autowidth(ws5)

    # ── Sheet 6: anomalies
    ws6 = wb.create_sheet("anomalies")
    if anomalies_all:
        cols = list(anomalies_all[0].keys())
        header_row(ws6, cols)
        for a in anomalies_all[:500]:
            ws6.append([str(a.get(c, "")) for c in cols])
    autowidth(ws6)

    path = reports_root / "excel_vs_crm_reconciliation.xlsx"
    wb.save(str(path))
    log.info(f"  Reconciliation XLSX -> {path}")
    return path


# ──── ЭТАП 7: ВЕРИФИКАЦИЯ BASELINE ────────────────────────────────────────────

def verify_baseline(total_crm_now):
    bl_path = OUT / "data" / "baseline_debt.json"
    if not bl_path.exists():
        log.warning("baseline_debt.json not found -- skip verification")
        return
    with open(bl_path, encoding="utf-8") as f:
        baseline = json.load(f)
    bl_total = baseline["total_debt"]
    diff = total_crm_now - bl_total
    log.info(f"\n{'='*60}")
    log.info("ВЕРИФИКАЦИЯ BASELINE")
    log.info(f"  Baseline (start of run): {bl_total:>18,.2f} сум")
    log.info(f"  CRM now (end of run):    {total_crm_now:>18,.2f} сум")
    log.info(f"  Diff:                    {diff:>18,.2f} сум")
    if abs(diff) < TOLERANCE:
        log.info("  OK -- долг CRM не изменился (скрипт только читал данные)")
    else:
        log.warning(f"  ВНИМАНИЕ: долг изменился на {diff:,.2f} сум!")
        log.warning("    Это изменение могло произойти из-за сторонних операций.")
    log.info(f"{'='*60}")


# ──── FUZZY MATCHING REPORT ───────────────────────────────────────────────────

def run_fuzzy_report(orders):
    all_norms = list(set(o["client_norm"] for o in orders if o["client_norm"]))
    suggestions = fuzzy_match_clients(all_norms)
    if suggestions:
        log.info(f"\n  FUZZY MATCH -- {len(suggestions)} предложений по слиянию имён:")
        for a, b, dist in suggestions[:30]:
            log.info(f"    '{a}' <-> '{b}'  (distance={dist})")
        if len(suggestions) > 30:
            log.info(f"    ... и ещё {len(suggestions)-30} пар")
    else:
        log.info("  Fuzzy match: совпадений не найдено")
    return suggestions


# ──── MAIN ────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 70)
    log.info(f"EXCEL AUDIT FULL v2.0 -- START  {TS}")
    log.info("=" * 70)
    log.info("(Dry-run: скрипт только читает Excel и БД, ничего не изменяет)\n")

    # ── Шаг 0: Бэкап уже сделан, baseline уже записан
    bl_path = OUT / "data" / "baseline_debt.json"
    with open(bl_path, encoding="utf-8") as f:
        baseline = json.load(f)
    log.info(f"Baseline loaded: {baseline['total_debt']:,.2f} сум "
             f"({baseline['timestamp']})")

    # ── Шаг 1: Скан
    log.info("\n-- ШАГ 1: ПОЛНЫЙ СКАН EXCEL-ФАЙЛОВ --")
    all_scans = scan_all_files()
    total_rows = sum(len(s["raw_rows"]) for s in all_scans)
    total_anom = sum(len(s["anomalies"]) for s in all_scans)
    log.info(f"\n  Итого: {len(all_scans)} листов, {total_rows} строк, {total_anom} аномалий")

    log.info("\n  Листы:")
    for s in all_scans:
        log.info(f"    {s['file_key']:<15} / {s['sheet']:<25} "
                 f"rows={len(s['raw_rows']):>5}  anomalies={len(s['anomalies']):>3}")

    # ── Шаг 2: Нормализация
    log.info("\n-- ШАГ 2: НОРМАЛИЗАЦИЯ --")
    orders, payments_normalized = build_orders_and_payments(all_scans)
    log.info(f"  Orders:   {len(orders)}")
    log.info(f"  Payments: {len(payments_normalized)}")

    # Stats: prepay and credit
    n_credit = sum(1 for o in orders if o["is_debt_type"])
    n_prepay = sum(1 for o in orders if o["is_prepay_type"])
    tot_prepay_sum = sum(o["prepay_sum"] for o in orders)
    log.info(f"  Credit-type rows (к/н/к/п/к/ф/фт/фотих): {n_credit}")
    log.info(f"  Prepay-type rows (ПП):                    {n_prepay}")
    log.info(f"  Total prepay_sum detected:                {tot_prepay_sum:,.0f} сум")

    # ── Шаг 3: Фуззи-матчинг имён
    log.info("\n-- ШАГ 3: FUZZY MATCHING ИМЁН КЛИЕНТОВ --")
    fuzzy_suggestions = run_fuzzy_report(orders)

    # ── Шаг 4: Excel-агрегат долгов (с prepay/net_debt)
    log.info("\n-- ШАГ 4: АГРЕГАЦИЯ EXCEL-ДОЛГОВ (+ PrepaySum + NetDebt) --")
    excel_agg = aggregate_excel_debts(orders)
    total_excel_remaining = sum(ea["remaining"] for ea in excel_agg.values() if ea["remaining"] > 0)
    total_excel_prepay = sum(ea["prepay_sum"] for ea in excel_agg.values())
    total_excel_net = sum(max(0, ea["net_debt"]) for ea in excel_agg.values())
    n_debtors = len([ea for ea in excel_agg.values() if ea["remaining"] > 0])
    n_net_debtors = len([ea for ea in excel_agg.values() if ea["net_debt"] > 0])

    log.info(f"  Клиентов с remaining > 0: {n_debtors}")
    log.info(f"  Клиентов с net_debt > 0:  {n_net_debtors}")
    log.info(f"  Sum remaining:            {total_excel_remaining:,.2f} сум")
    log.info(f"  Sum prepay:               {total_excel_prepay:,.2f} сум")
    log.info(f"  Sum net_debt:             {total_excel_net:,.2f} сум")

    # ── Шаг 5: CRM-данные
    log.info("\n-- ШАГ 5: ЗАГРУЗКА ДАННЫХ CRM --")
    crm_clients, total_crm = load_crm_debts()
    log.info(f"  Клиентов с долгом в CRM: {len(crm_clients)}")
    log.info(f"  Суммарный долг CRM:      {total_crm:,.2f} сум")

    # ── Шаг 6: Реконсилиэйшн
    log.info("\n-- ШАГ 6: РЕКОНСИЛИЭЙШН EXCEL <-> CRM --")
    rec = reconcile(excel_agg, crm_clients, total_crm)
    log.info(f"  Глобальная разница (CRM - Excel remaining): {rec['global_diff']:,.2f} сум")
    log.info(f"  Расхождений > {TOLERANCE:,} сум: {len(rec['diffs'])}")
    log.info(f"  Только в CRM:      {len(rec['only_crm'])}")
    log.info(f"  Только в Excel:    {len(rec['only_excel'])}")

    log.info("\n  ТОП-20 расхождений:")
    log.info(f"  {'Клиент':<30} {'Remaining':>13} {'Prepay':>10} {'NetDebt':>13} {'CRM':>13} {'Diff':>13} {'Status'}")
    log.info("  " + "-" * 100)
    for d in rec["diffs"][:20]:
        nm = (d.get("company_name") or d["client_norm"])[:28]
        log.info(f"  {nm:<30} {d['excel_remaining']:>13,.0f} {d['excel_prepay']:>10,.0f} "
                 f"{d['excel_net_debt']:>13,.0f} {d['crm_debt']:>13,.0f} "
                 f"{d['diff']:>13,.0f}  {d['status']}")

    # ── Шаг 7: Экспорт
    log.info("\n-- ШАГ 7: ЭКСПОРТ АРТЕФАКТОВ --")
    all_anomalies = [a for s in all_scans for a in s["anomalies"]]
    export_csvs(all_scans, orders, payments_normalized, all_anomalies, rec, excel_agg)

    # Сохранить fuzzy suggestions
    fuzzy_path = OUT / "data" / "reports" / f"fuzzy_suggestions_{TS}.json"
    with open(fuzzy_path, "w", encoding="utf-8") as f:
        json.dump(fuzzy_suggestions, f, ensure_ascii=False, indent=2)
    log.info(f"  Fuzzy suggestions -> {fuzzy_path}")

    # ── Шаг 8: Верификация baseline
    log.info("\n-- ШАГ 8: ВЕРИФИКАЦИЯ BASELINE --")
    verify_baseline(total_crm)

    # ── Итог
    log.info("\n" + "=" * 70)
    log.info("ИТОГОВАЯ СВОДКА v2.0")
    log.info("=" * 70)
    log.info(f"  Файлов обработано:          {len(FILES)}")
    log.info(f"  Листов обработано:          {len(all_scans)}")
    log.info(f"  Строк (raw):                {total_rows}")
    log.info(f"  Orders (нормализованных):   {len(orders)}")
    log.info(f"  Payments (нормализованных): {len(payments_normalized)}")
    log.info(f"  Аномалий:                   {total_anom}")
    log.info(f"  Fuzzy пар для проверки:     {len(fuzzy_suggestions)}")
    log.info(f"  -------")
    log.info(f"  Excel remaining:            {total_excel_remaining:>18,.2f} сум")
    log.info(f"  Excel prepay_sum:           {total_excel_prepay:>18,.2f} сум")
    log.info(f"  Excel net_debt:             {total_excel_net:>18,.2f} сум")
    log.info(f"  CRM debt (current):         {total_crm:>18,.2f} сум")
    log.info(f"  Разница (CRM-remaining):    {rec['global_diff']:>18,.2f} сум")
    log.info(f"\n  Лог: {log_path}")
    log.info(f"  Отчёт: mnt/data/reports/excel_vs_crm_reconciliation.xlsx")
    log.info(f"  Per-client: mnt/data/reports/per_client_agg_{TS}.csv")
    log.info("=" * 70)

    return {
        "scans": len(all_scans),
        "rows": total_rows,
        "orders": len(orders),
        "payments": len(payments_normalized),
        "anomalies": total_anom,
        "total_excel_remaining": total_excel_remaining,
        "total_excel_prepay": total_excel_prepay,
        "total_excel_net": total_excel_net,
        "total_crm": total_crm,
        "diffs": len(rec["diffs"]),
        "only_crm": len(rec["only_crm"]),
        "only_excel": len(rec["only_excel"]),
        "fuzzy": len(fuzzy_suggestions),
    }


if __name__ == "__main__":
    result = main()
