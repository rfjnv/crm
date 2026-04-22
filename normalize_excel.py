"""
Финансовый Excel нормализатор и аудитор
========================================
Версия: 1.0 | Дата: 2026-03-09

Извлекает данные из Excel-файлов (множество листов/месяцев),
нормализует даты, суммы, способы оплаты, сроки оплаты и остатки (долги).
Выводит каноническую таблицу + summary report.
"""

import os
import re
import sys
import json
import logging
from pathlib import Path
from datetime import datetime, timedelta
from calendar import monthrange
from collections import defaultdict

import openpyxl
import pandas as pd
import numpy as np

# ─── SETTINGS ──────────────────────────────────────────────────────────────────
ROOT = Path(r"c:\Users\Noutbuk savdosi\CRM")
TODAY = datetime(2026, 3, 9)

FILES = [
    ROOT / "07.03.2026.xlsx",                # newest 2026 data (replaces 05.03.2026.xlsx)
    ROOT / "29.12.2025.xlsx",
    ROOT / "26.12.2024.xlsx",
]

OUTPUT_CSV   = ROOT / "mnt" / "data" / "normalized_output.csv"
SUMMARY_JSON = ROOT / "mnt" / "data" / "summary_report.json"
SUMMARY_TXT  = ROOT / "mnt" / "data" / "summary_report.txt"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(ROOT / "mnt" / "data" / "normalize.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ─── MONTH NAME MAP ───────────────────────────────────────────────────────────
MONTH_NAMES_RU = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}

# ─── PAYMENT METHOD MAPPING ──────────────────────────────────────────────────
PAYMENT_METHOD_MAP = {
    "к":    "долг",
    "н":    "наличными",
    "п":    "перечисление",
    "ф":    "Фотих",
    "т":    "терминал",
    "пп":   "передоплата",
    "н/к":  "долг с наличными",
    "п/к":  "перечисление в долг",
}

# Extended codes that mean "credit/debt"
CREDIT_CODES = {"к", "н/к", "п/к", "пк/к", "ф", "фт", "фотих"}
PREPAY_CODES = {"пп"}

# Substring-based fallback detection
PAYMENT_SUBSTRINGS = {
    "наличн":   "наличными",
    "терминал":  "терминал",
    "перечисл":  "перечисление",
    "передоплат": "передоплата",
    "долг":      "долг",
    "qr":        "QR CODE",
    "клик":      "клик",
}


# ─── COLUMN MAPPING PER (year, month) ────────────────────────────────────────
# This is built from examining the actual Excel structure.
# Keys: (year, month_num)
# Values: dict with 0-based column indices
#   date_col, client_col, carry_col, product_col, qty_col, unit_col,
#   price_col, amount_col, nkp_col, deadline_col, payment_cols (dict),
#   end_balance_col, payment_date_col

def _col_idx(letter):
    """Convert Excel column letter to 0-based index."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1

# Layout type 1: Jan 2026, and 2025 sheets (28 cols variant)
# A=date, B=client, C=carry, E=product, F=qty, G=unit, H=price, I=amount,
# J=nkp, K=contract, L=payment_cash_total, M=month_cash,
# O=per_total, P=month_per, R=qr_total, S=month_qr,
# U=click_total, V=month_click, X=term_total, Y=month_term,
# AA=end_balance, AB=payment_date
LAYOUT_28 = {
    "date": _col_idx("A"),
    "client": _col_idx("B"),
    "carry": _col_idx("C"),
    "manager": _col_idx("D"),
    "product": _col_idx("E"),
    "qty": _col_idx("F"),
    "unit": _col_idx("G"),
    "price": _col_idx("H"),
    "amount": _col_idx("I"),
    "nkp": _col_idx("J"),
    "deadline": _col_idx("K"),
    "cash_total": _col_idx("L"),
    "cash_month": _col_idx("M"),
    "per_total": _col_idx("O"),
    "per_month": _col_idx("P"),
    "qr_total": _col_idx("R"),
    "qr_month": _col_idx("S"),
    "click_total": _col_idx("U"),
    "click_month": _col_idx("V"),
    "term_total": _col_idx("X"),
    "term_month": _col_idx("Y"),
    "end_balance": _col_idx("AA"),
    "payment_date": _col_idx("AB"),
}

# Layout type 2: Feb/Mar 2026 (29 cols variant)
# A=date, B=client, C=carry, E=product, F=qty, G=unit, H=price, I=amount,
# J=nkp, K=deadline, L=contract, M=payment_cash_total, N=month_cash,
# P=per_total, Q=month_per, S=qr_total, T=month_qr,
# V=click_total, W=month_click, Y=term_total, Z=month_term,
# AB=end_balance, AC=payment_date
LAYOUT_29 = {
    "date": _col_idx("A"),
    "client": _col_idx("B"),
    "carry": _col_idx("C"),
    "manager": _col_idx("D"),
    "product": _col_idx("E"),
    "qty": _col_idx("F"),
    "unit": _col_idx("G"),
    "price": _col_idx("H"),
    "amount": _col_idx("I"),
    "nkp": _col_idx("J"),
    "deadline": _col_idx("K"),
    "contract": _col_idx("L"),
    "cash_total": _col_idx("M"),
    "cash_month": _col_idx("N"),
    "per_total": _col_idx("P"),
    "per_month": _col_idx("Q"),
    "qr_total": _col_idx("S"),
    "qr_month": _col_idx("T"),
    "click_total": _col_idx("V"),
    "click_month": _col_idx("W"),
    "term_total": _col_idx("Y"),
    "term_month": _col_idx("Z"),
    "end_balance": _col_idx("AB"),
    "payment_date": _col_idx("AC"),
}

# 2024 layout (30+ cols, slightly different)
LAYOUT_2024_JAN = {
    "date": _col_idx("A"),
    "client": _col_idx("B"),
    "carry": _col_idx("C"),
    "manager": _col_idx("D"),
    "product": _col_idx("F"),
    "qty": _col_idx("G"),
    "unit": _col_idx("H"),
    "price": _col_idx("I"),
    "amount": _col_idx("J"),
    "nkp": _col_idx("M"),
    "deadline": _col_idx("N"),
    "cash_total": _col_idx("O"),
    "cash_month": _col_idx("P"),
    "per_total": _col_idx("R"),
    "per_month": _col_idx("S"),
    "qr_total": _col_idx("U"),
    "qr_month": _col_idx("V"),
    "click_total": _col_idx("X"),
    "click_month": _col_idx("Y"),
    "term_total": _col_idx("AA"),
    "term_month": _col_idx("AB"),
    "end_balance": _col_idx("AE"),
    "payment_date": _col_idx("AG"),
}

LAYOUT_2024 = {
    "date": _col_idx("A"),
    "client": _col_idx("B"),
    "carry": _col_idx("C"),
    "manager": _col_idx("D"),
    "product": _col_idx("E"),
    "qty": _col_idx("F"),
    "unit": _col_idx("G"),
    "price": _col_idx("H"),
    "amount": _col_idx("I"),
    "nkp": _col_idx("L"),
    "deadline": _col_idx("M"),
    "cash_total": _col_idx("N"),
    "cash_month": _col_idx("O"),
    "per_total": _col_idx("Q"),
    "per_month": _col_idx("R"),
    "qr_total": _col_idx("T"),
    "qr_month": _col_idx("U"),
    "click_total": _col_idx("W"),
    "click_month": _col_idx("X"),
    "term_total": _col_idx("Z"),
    "term_month": _col_idx("AA"),
    "end_balance": _col_idx("AC"),
    "payment_date": _col_idx("AD"),
}


def detect_layout(ws, sheet_name, file_name):
    """Detect which column layout applies to this sheet."""
    max_col = ws.max_column or 28
    # Read header row 1
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        for c in row:
            if c.value:
                headers[c.column_letter] = str(c.value).strip().lower()

    sheet_lower = sheet_name.lower()
    # Determine year from sheet name
    year = None
    for token in sheet_lower.split():
        if token.isdigit() and len(token) == 4:
            year = int(token)
            break

    if year and year <= 2024:
        # 2024 files: Jan 2024 has 33 cols, others have 30
        if max_col >= 33:
            return LAYOUT_2024_JAN
        return LAYOUT_2024

    # 2025-2026: check if layout is 28 or 29 columns
    # Key distinguisher: in layout_28, col K is "договор"/"Оплата" (no separate deadline/contract)
    #                     in layout_29, col K is "срок опалата", col L is "договор номер"
    k_header = headers.get("K", "")
    if "срок" in k_header or "опалата" in k_header or "оплат" in k_header:
        return LAYOUT_29
    if max_col >= 29:
        # Check row 1 for AC column (число)
        ac_header = headers.get("AC", "")
        if ac_header:
            return LAYOUT_29
    return LAYOUT_28


def parse_sheet_period(sheet_name):
    """Extract (year, month_num) from sheet name like 'февраль 2026'."""
    parts = sheet_name.lower().strip().split()
    month_num = None
    year = None
    for p in parts:
        if p in MONTH_NAMES_RU:
            month_num = MONTH_NAMES_RU[p]
        elif p.isdigit() and len(p) == 4:
            year = int(p)
    return year, month_num


# ─── NORMALIZATION HELPERS ───────────────────────────────────────────────────
def normalize_amount(val):
    """Convert a value to numeric amount."""
    if val is None:
        return np.nan
    if isinstance(val, (int, float)):
        if isinstance(val, float) and abs(val) < 1e-6:
            return 0.0
        return float(val)
    s = str(val).strip()
    s = re.sub(r'[^0-9.\-]', '', s.replace(',', '.').replace(' ', ''))
    if not s or s == '-':
        return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan


def parse_date(val):
    """Parse a date value from Excel cell."""
    if val is None:
        return pd.NaT
    if isinstance(val, datetime):
        return pd.Timestamp(val)
    s = str(val).strip()
    # Try DD.MM.YYYY or DD.MM.YY
    for fmt in ["%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"]:
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except ValueError:
            continue
    # Try pandas parser as last resort
    try:
        return pd.Timestamp(pd.to_datetime(s, dayfirst=True))
    except Exception:
        return pd.NaT


def extract_deadline_from_text(text):
    """Extract payment deadline date from text like 'оплата 05.12.25' or 'дог №847'."""
    if not text:
        return pd.NaT, 0, ""
    s = str(text).strip().lower()

    # Check if it's just a contract number (дог №...)
    if re.match(r'^(дог|п)\s*[№#]?\s*\d+$', s) or re.match(r'^\d{3,10}$', s):
        return pd.NaT, 0, "contract_number_not_date"

    # Check for "оплата DD.MM.YY" pattern
    m = re.search(r'оплат[аы]?\s*(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            dt = pd.Timestamp(datetime(y, mo, d))
            confidence = 0.8
            return dt, confidence, f"extracted_from_text: '{text}'"
        except ValueError:
            pass

    # Try parsing date patterns without 'оплата' keyword
    m = re.search(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            dt = pd.Timestamp(datetime(y, mo, d))
            confidence = 0.5
            return dt, confidence, f"date_extracted_no_keyword: '{text}'"
        except ValueError:
            pass

    return pd.NaT, 0, f"unrecognized: '{text}'"


def normalize_payment_method(raw, description=""):
    """Normalize payment method code."""
    if raw is None or (isinstance(raw, float) and np.isnan(raw)):
        return None, 0, False, "no_payment_method"

    code = str(raw).strip().lower()
    needs_review = False
    confidence = 1.0
    log_msg = f"raw='{raw}'"

    # Direct mapping
    if code in PAYMENT_METHOD_MAP:
        normalized = PAYMENT_METHOD_MAP[code]
        if code == "ф":
            needs_review = True
            log_msg += " | ф→Фотих needs_review"
        return normalized, confidence, needs_review, log_msg

    # Extended codes
    if code in ("фт", "фотих"):
        return "Фотих", 1.0, True, log_msg + " | Фотих needs_review"
    if code in ("пк/к",):
        return "перечисление в долг", 1.0, False, log_msg
    if code in ("нал", "нкл"):
        return "наличными", 0.8, False, log_msg + " | inferred from variant"
    if code in ("б/н",):
        return "перечисление", 0.8, False, log_msg + " | inferred from б/н"

    # Substring search in raw + description
    search_text = (code + " " + str(description or "")).lower()
    for substr, norm_val in PAYMENT_SUBSTRINGS.items():
        if substr in search_text:
            return norm_val, 0.6, False, log_msg + f" | inferred from substring '{substr}'"

    return None, 0.0, True, log_msg + " | unrecognized"


def compute_payment_date(indicator_val, date_created, sheet_year, sheet_month):
    """
    Process the 'число' column value.
    If it's a datetime → use directly (confidence=1).
    If it's an integer 1-31 → construct date (confidence=0.5).
    """
    if indicator_val is None:
        return pd.NaT, 0.0, "no_indicator"

    # Already a datetime
    if isinstance(indicator_val, datetime):
        return pd.Timestamp(indicator_val), 1.0, "direct_datetime"

    # Try as integer (day number)
    try:
        day = int(float(str(indicator_val)))
    except (ValueError, TypeError):
        # Try parsing as date string
        dt = parse_date(indicator_val)
        if pd.notna(dt):
            return dt, 0.8, f"parsed_from_string: '{indicator_val}'"
        return pd.NaT, 0.0, f"unparseable: '{indicator_val}'"

    if 1 <= day <= 31:
        # Construct date from day + context month/year
        ref_year = sheet_year
        ref_month = sheet_month

        # If date_created is available, use its month/year
        if pd.notna(date_created):
            try:
                ref_year = date_created.year
                ref_month = date_created.month
            except Exception:
                pass

        # Try constructing the date
        try:
            max_day = monthrange(ref_year, ref_month)[1]
            actual_day = min(day, max_day)
            candidate = pd.Timestamp(datetime(ref_year, ref_month, actual_day))

            # If candidate < date_created → next month
            if pd.notna(date_created) and candidate < date_created:
                if ref_month == 12:
                    ref_month = 1
                    ref_year += 1
                else:
                    ref_month += 1
                max_day = monthrange(ref_year, ref_month)[1]
                actual_day = min(day, max_day)
                candidate = pd.Timestamp(datetime(ref_year, ref_month, actual_day))
                return candidate, 0.5, f"day={day} applied to next month {ref_year}-{ref_month:02d}"

            return candidate, 0.5, f"day={day} applied to {ref_year}-{ref_month:02d}"
        except Exception as e:
            return pd.NaT, 0.0, f"day={day} construction failed: {e}"

    return pd.NaT, 0.0, f"day_out_of_range: {indicator_val}"


def compute_total_payments(row_data, layout):
    """Sum all payment columns for a row to get total paid amount."""
    total = 0.0
    payment_keys = ["cash_month", "per_month", "qr_month", "click_month", "term_month"]
    for key in payment_keys:
        idx = layout.get(key)
        if idx is not None and idx < len(row_data):
            val = normalize_amount(row_data[idx])
            if not np.isnan(val):
                total += val
    return total


# ─── MAIN PROCESSING ────────────────────────────────────────────────────────
def process_sheet(ws, sheet_name, file_path, layout, sheet_year, sheet_month):
    """Process a single Excel sheet and return list of canonical row dicts."""
    rows = []
    source_file = file_path.name
    data_start_row = 4  # rows 1-3 are headers

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True),
                                    start=data_start_row):
        row_data = list(row)
        proc_log = []

        # ── Extract raw values ──
        def get_val(key):
            idx = layout.get(key)
            if idx is not None and idx < len(row_data):
                return row_data[idx]
            return None

        raw_date = get_val("date")
        raw_client = get_val("client")
        raw_carry = get_val("carry")
        raw_amount = get_val("amount")
        raw_nkp = get_val("nkp")
        raw_deadline = get_val("deadline")
        raw_end_balance = get_val("end_balance")
        raw_payment_date = get_val("payment_date")
        raw_product = get_val("product")

        # Skip completely empty rows
        if raw_client is None and raw_amount is None and raw_carry is None and raw_end_balance is None:
            continue

        # Skip if client is empty and no financial data
        has_financial = (
            raw_amount is not None
            or raw_carry is not None
            or raw_end_balance is not None
            or raw_nkp is not None
        )
        if raw_client is None and not has_financial:
            continue

        needs_review = False

        # ── Normalize date ──
        date_created = parse_date(raw_date)
        if pd.isna(date_created) and raw_date is not None:
            proc_log.append(f"date_parse_failed: '{raw_date}'")
            needs_review = True

        # ── Normalize client ──
        client = str(raw_client).strip() if raw_client else ""

        # ── Normalize amounts ──
        carry_debt = normalize_amount(raw_carry)
        amount = normalize_amount(raw_amount)
        end_balance = normalize_amount(raw_end_balance)

        if np.isnan(amount) and raw_amount is not None:
            proc_log.append(f"amount_parse_failed: '{raw_amount}'")
            needs_review = True

        # Compute sale amount (from I column) — this is the sale total
        sale_amount = amount

        # Compute total payments from payment sub-columns
        total_paid = compute_total_payments(row_data, layout)

        # ── Payment method ──
        pm_normalized, pm_confidence, pm_review, pm_log = normalize_payment_method(
            raw_nkp, raw_product
        )
        proc_log.append(f"pm: {pm_log}")
        if pm_review:
            needs_review = True

        # ── Payment date (число column) ──
        payment_date, pd_confidence, pd_log = compute_payment_date(
            raw_payment_date, date_created, sheet_year, sheet_month
        )
        proc_log.append(f"payment_date: {pd_log}")

        # ── Payment deadline (срок оплаты from K) ──
        deadline = pd.NaT
        deadline_confidence = 0
        deadline_log = ""
        if raw_deadline is not None:
            if isinstance(raw_deadline, datetime):
                deadline = pd.Timestamp(raw_deadline)
                deadline_confidence = 1.0
                deadline_log = "direct_datetime"
            else:
                deadline, deadline_confidence, deadline_log = extract_deadline_from_text(raw_deadline)
        proc_log.append(f"deadline: {deadline_log}")

        # ── Payment status ──
        payment_status = ""
        days_overdue = np.nan
        if pd.notna(deadline):
            if pd.notna(payment_date):
                if payment_date <= deadline:
                    payment_status = "on_time"
                    days_overdue = 0
                else:
                    payment_status = "overdue"
                    days_overdue = (payment_date - deadline).days
            else:
                # No payment date
                if pd.Timestamp(TODAY) > deadline:
                    payment_status = "overdue_unpaid"
                    days_overdue = (pd.Timestamp(TODAY) - deadline).days
                else:
                    payment_status = "pending"
                    days_overdue = 0
        elif pd.notna(payment_date):
            payment_status = "paid_no_deadline"
            proc_log.append("no_deadline_for_status")

        # ── Balance check ──
        if not np.isnan(carry_debt) and not np.isnan(sale_amount) and not np.isnan(end_balance):
            expected_end = carry_debt + sale_amount - total_paid
            diff = abs(expected_end - end_balance)
            if diff > max(abs(end_balance) * 0.005, 1000):
                proc_log.append(
                    f"balance_check_MISMATCH: carry({carry_debt})+sale({sale_amount})-paid({total_paid})="
                    f"{expected_end} vs end_balance({end_balance}), diff={diff:.0f}"
                )
                needs_review = True
            else:
                proc_log.append("balance_check_OK")

        # ── Payment > sale check ──
        if total_paid > 0 and not np.isnan(sale_amount) and sale_amount > 0:
            if total_paid > sale_amount * 1.01:  # 1% tolerance
                proc_log.append(
                    f"payment_exceeds_sale: paid={total_paid} > sale={sale_amount}"
                )
                needs_review = True

        # ── Build canonical row ──
        rows.append({
            "source_file": source_file,
            "sheet": sheet_name,
            "sheet_year": sheet_year,
            "sheet_month": sheet_month,
            "row_number": row_idx,
            "client": client,
            "order_id": "",
            "date": date_created.strftime("%Y-%m-%d") if pd.notna(date_created) else "",
            "payment_date": payment_date.strftime("%Y-%m-%d") if pd.notna(payment_date) else "",
            "payment_date_confidence": pd_confidence,
            "amount": sale_amount if not np.isnan(sale_amount) else "",
            "total_paid": total_paid,
            "payment_method_raw": str(raw_nkp) if raw_nkp is not None else "",
            "payment_method_normalized": pm_normalized or "",
            "payment_method_confidence": pm_confidence,
            "carry_debt": carry_debt if not np.isnan(carry_debt) else "",
            "end_balance": end_balance if not np.isnan(end_balance) else "",
            "payment_deadline": deadline.strftime("%Y-%m-%d") if pd.notna(deadline) else "",
            "payment_deadline_confidence": deadline_confidence,
            "payment_status": payment_status,
            "days_overdue": int(days_overdue) if not np.isnan(days_overdue) else "",
            "processing_log": " | ".join(proc_log),
            "needs_review": needs_review,
        })

    return rows


def process_file(file_path):
    """Process an entire Excel file (all sheets)."""
    log.info(f"Processing file: {file_path}")
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        # Skip non-month sheets (like "итого", "лист1" etc.)
        sheet_year, sheet_month = parse_sheet_period(sheet_name)
        if sheet_year is None or sheet_month is None:
            log.info(f"  Skipping sheet '{sheet_name}' (no year/month detected)")
            continue

        log.info(f"  Processing sheet: {sheet_name} -> {sheet_year}-{sheet_month:02d}")
        ws = wb[sheet_name]
        layout = detect_layout(ws, sheet_name, file_path.name)
        rows = process_sheet(ws, sheet_name, file_path, layout, sheet_year, sheet_month)
        log.info(f"    -> {len(rows)} rows extracted")
        all_rows.extend(rows)

    wb.close()
    return all_rows


def generate_carryover_rows(df):
    """
    For each client with carry_debt > 0 at end of month,
    create a carryover row for the first day of the next month.
    """
    carry_rows = []
    # Group by client + sheet year/month, take last carry_debt
    for (client, year, month), group in df.groupby(["client", "sheet_year", "sheet_month"]):
        if not client:
            continue
        # Get the last end_balance for this client this month
        last_end = group["end_balance"].dropna()
        if last_end.empty:
            continue
        last_val = last_end.iloc[-1]
        if isinstance(last_val, str) and last_val == "":
            continue
        last_val = float(last_val)
        if last_val > 0:
            # Next month
            if month == 12:
                next_year, next_month = year + 1, 1
            else:
                next_year, next_month = year, month + 1

            carry_rows.append({
                "source_file": "carryover",
                "sheet": f"carryover_{year}-{month:02d}→{next_year}-{next_month:02d}",
                "sheet_year": next_year,
                "sheet_month": next_month,
                "row_number": 0,
                "client": client,
                "order_id": "",
                "date": f"{next_year}-{next_month:02d}-01",
                "payment_date": "",
                "payment_date_confidence": 0,
                "amount": 0,
                "total_paid": 0,
                "payment_method_raw": "",
                "payment_method_normalized": "долг",
                "payment_method_confidence": 1.0,
                "carry_debt": last_val,
                "end_balance": "",
                "payment_deadline": "",
                "payment_deadline_confidence": 0,
                "payment_status": "",
                "days_overdue": "",
                "processing_log": f"carryover applied from {year}-{month:02d}, amount={last_val}",
                "needs_review": False,
            })

    return carry_rows


def generate_summary(df):
    """Generate summary report."""
    summary = {}

    # Latest report date (from real data, not carryover)
    real_df = df[df["source_file"] != "carryover"].copy()
    dates = pd.to_datetime(real_df["date"], errors="coerce")
    latest_date = dates.max()
    summary["latest_report_date"] = str(latest_date.date()) if pd.notna(latest_date) else "unknown"

    # Latest period: not used below, remove stale code

    # Total debt: sum end_balance for latest month rows
    # Per spec: use end_month_balance from latest report period
    # IMPORTANT: exclude carryover rows when determining latest period
    real = df[df["source_file"] != "carryover"]
    latest_year = real["sheet_year"].max()
    latest_month = real.loc[real["sheet_year"] == latest_year, "sheet_month"].max()
    latest_mask = (df["sheet_year"] == latest_year) & (df["sheet_month"] == latest_month)
    latest_data = df.loc[latest_mask & (df["source_file"] != "carryover")]

    end_balances = pd.to_numeric(latest_data["end_balance"], errors="coerce")
    total_debt = end_balances.sum()
    positive_debt = end_balances[end_balances > 0].sum()
    negative_debt = end_balances[end_balances < 0].sum()

    summary["latest_period"] = f"{latest_year}-{latest_month:02d}"
    summary["total_end_balance"] = float(total_debt)
    summary["total_positive_debt"] = float(positive_debt)
    summary["total_negative_balance_prepaid"] = float(negative_debt)
    summary["rows_in_latest_period"] = int(latest_mask.sum())
    summary["total_rows"] = len(df)

    # needs_review stats
    review_mask = df["needs_review"] == True
    summary["needs_review_count"] = int(review_mask.sum())

    # Top 10 needs_review examples
    review_examples = df.loc[review_mask].head(10)
    summary["needs_review_top10"] = []
    for _, r in review_examples.iterrows():
        summary["needs_review_top10"].append({
            "file": r["source_file"],
            "sheet": r["sheet"],
            "row": int(r["row_number"]),
            "client": r["client"],
            "log": r["processing_log"][:200],
        })

    # Per-client debt (latest period)
    per_client = (
        latest_data.assign(eb=end_balances)
        .groupby("client")["eb"]
        .sum()
        .sort_values(ascending=False)
    )
    summary["per_client_debt_top20"] = {
        k: float(v) for k, v in per_client.head(20).items() if v != 0
    }

    # Payment method distribution
    pm_dist = df["payment_method_normalized"].value_counts().to_dict()
    summary["payment_method_distribution"] = {str(k): int(v) for k, v in pm_dist.items()}

    # Carryover stats
    carry_mask = df["source_file"] == "carryover"
    summary["carryover_rows_created"] = int(carry_mask.sum())

    # Payment status distribution (latest period)
    status_dist = latest_data["payment_status"].value_counts().to_dict()
    summary["payment_status_distribution"] = {str(k): int(v) for k, v in status_dist.items()}

    # Overdue summary
    overdue = latest_data[latest_data["payment_status"].isin(["overdue", "overdue_unpaid"])]
    summary["overdue_count"] = len(overdue)
    if not overdue.empty:
        days = pd.to_numeric(overdue["days_overdue"], errors="coerce")
        summary["overdue_avg_days"] = float(days.mean()) if days.notna().any() else 0
        summary["overdue_max_days"] = float(days.max()) if days.notna().any() else 0

    return summary


def main():
    log.info("=" * 70)
    log.info("EXCEL NORMALIZER STARTED")
    log.info("=" * 70)

    all_rows = []
    for fp in FILES:
        if fp.exists():
            file_rows = process_file(fp)
            all_rows.extend(file_rows)
        else:
            log.warning(f"File not found: {fp}")

    log.info(f"Total rows extracted: {len(all_rows)}")

    # Convert to DataFrame
    df = pd.DataFrame(all_rows)
    if df.empty:
        log.error("No data extracted!")
        return

    # Generate carryover rows
    carry_rows = generate_carryover_rows(df)
    log.info(f"Carryover rows generated: {len(carry_rows)}")
    if carry_rows:
        df_carry = pd.DataFrame(carry_rows)
        df = pd.concat([df, df_carry], ignore_index=True)

    # Sort by date, then by sheet
    df = df.sort_values(
        ["sheet_year", "sheet_month", "client", "row_number"],
        na_position="last",
    ).reset_index(drop=True)

    # Generate summary
    summary = generate_summary(df)

    # Ensure output directory exists
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)

    # Save CSV
    df.to_csv(str(OUTPUT_CSV), index=False, encoding="utf-8-sig")
    log.info(f"Output CSV saved: {OUTPUT_CSV}")

    # Save summary
    with open(str(SUMMARY_JSON), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    log.info(f"Summary report saved: {SUMMARY_JSON}")

    # Print summary
    lines = []
    def out(s=""):
        lines.append(s)
        print(s)

    out("\n" + "=" * 70)
    out("PRODUCTION RUN - SUMMARY REPORT")
    out(f"Run date: {TODAY.strftime('%Y-%m-%d')}")
    out("=" * 70)
    out(f"Latest period:            {summary['latest_period']}")
    out(f"Latest report date:       {summary['latest_report_date']}")
    out(f"Total rows processed:     {summary['total_rows']}")
    out(f"Rows in latest period:    {summary['rows_in_latest_period']}")
    out("")
    out("--- ДОЛГ ---")
    out(f"  Валовой долг (нам должны):    {summary['total_positive_debt']:>20,.0f}")
    out(f"  Предоплаты (переплаты):       {summary['total_negative_balance_prepaid']:>20,.0f}")
    out(f"                                {'=' * 20}")
    out(f"  ЧИСТЫЙ ДОЛГ:                  {summary['total_end_balance']:>20,.0f}")
    out("")
    out(f"Needs review:             {summary['needs_review_count']}")
    out(f"Carryover rows:           {summary['carryover_rows_created']}")
    out(f"Overdue count:            {summary.get('overdue_count', 0)}")
    if summary.get('overdue_avg_days'):
        out(f"Overdue avg days:         {summary['overdue_avg_days']:.0f}")
    if summary.get('overdue_max_days'):
        out(f"Overdue max days:         {summary['overdue_max_days']:.0f}")

    out("\nTop 20 Per-Client Debt (latest period):")
    for client, debt in summary.get("per_client_debt_top20", {}).items():
        if debt > 0:
            out(f"  {client:40s} {debt:>15,.0f}")

    out("\nPayment Method Distribution:")
    for method, count in summary.get("payment_method_distribution", {}).items():
        out(f"  {method or '(empty)':30s} {count:>6d}")

    out("\nPayment Status (latest period):")
    for status, count in summary.get("payment_status_distribution", {}).items():
        out(f"  {status or '(no deadline info)':30s} {count:>6d}")

    if summary.get("needs_review_top10"):
        out("\nTop 10 Needs Review:")
        for ex in summary["needs_review_top10"]:
            out(f"  [{ex['file']}:{ex['sheet']}:row {ex['row']}] {ex['client']}")
            out(f"    {ex['log'][:120]}")

    out("\n" + "=" * 70)
    out("OUTPUT FILES:")
    out(f"  CSV:     {OUTPUT_CSV}")
    out(f"  JSON:    {SUMMARY_JSON}")
    out(f"  TXT:     {SUMMARY_TXT}")
    out("=" * 70)
    out("DONE")
    out("=" * 70)

    # Save text summary
    with open(str(SUMMARY_TXT), "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    main()
