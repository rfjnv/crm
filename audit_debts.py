"""
Финансовый аудит долгов клиентов по Excel CRM отчёту типографии.
Источник: 03.03.2026.xlsx
Правило: для каждого клиента берём данные из ПОСЛЕДНЕГО листа, где он присутствует.
Колонка остатка: AB(28) для февраль/март, AA(27) для январь.
Не суммируем остатки разных месяцев.
"""
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\Noutbuk savdosi\CRM\03.03.2026.xlsx', data_only=True)

# Конфигурация листов: от старого к новому
# balance_col — колонка с остатком долга
sheets_config = [
    {'name': 'январь 2026',  'balance_col': 27, 'label': 'Ост на 31.01.2026 (кол AA)'},
    {'name': 'февраль 2026', 'balance_col': 28, 'label': 'Ост на 28.02.2026 (кол AB)'},
    {'name': 'март 2026',    'balance_col': 28, 'label': 'Ост на 31.03.2026 (кол AB)'},
]

CLIENT_COL = 2  # колонка B — название фирмы
DATA_START_ROW = 4  # данные начинаются с 4-й строки

def normalize(name):
    """Нормализация имени клиента для группировки."""
    if not name:
        return None
    s = str(name).strip().lower()
    s = s.replace('ё', 'е')
    # убираем лишние пробелы
    s = ' '.join(s.split())
    return s if s else None

# ============================================================
# Шаг 1: Собираем данные по каждому листу отдельно.
# Для каждого клиента сохраняем строки ТОЛЬКО из последнего листа.
# ============================================================

# Словарь: normalized_name -> {sheet_index, rows: [{balance, raw_name, row_num}]}
client_data = {}

for sheet_idx, cfg in enumerate(sheets_config):
    sheet_name = cfg['name']
    balance_col = cfg['balance_col']

    if sheet_name not in wb.sheetnames:
        print(f"ПРЕДУПРЕЖДЕНИЕ: Лист '{sheet_name}' не найден!")
        continue

    ws = wb[sheet_name]

    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        raw_name = ws.cell(row=row_num, column=CLIENT_COL).value
        norm = normalize(raw_name)
        if not norm:
            continue

        balance_val = ws.cell(row=row_num, column=balance_col).value

        # Пропускаем строки без числового остатка
        if balance_val is None:
            continue
        try:
            balance = float(balance_val)
        except (ValueError, TypeError):
            continue

        # Если клиент уже есть от более раннего листа — перезаписываем
        # Если от того же листа — добавляем строку
        if norm not in client_data or client_data[norm]['sheet_idx'] < sheet_idx:
            client_data[norm] = {
                'sheet_idx': sheet_idx,
                'sheet_name': sheet_name,
                'raw_name': str(raw_name).strip(),
                'rows': [{'balance': balance, 'row': row_num}],
            }
        elif client_data[norm]['sheet_idx'] == sheet_idx:
            client_data[norm]['rows'].append({'balance': balance, 'row': row_num})

# ============================================================
# Шаг 2: Суммируем долг по каждому клиенту (внутри одного листа).
# ============================================================

results = []
for norm, data in client_data.items():
    total_debt = sum(r['balance'] for r in data['rows'])
    row_count = len(data['rows'])
    results.append({
        'name': data['raw_name'],
        'norm': norm,
        'debt': total_debt,
        'rows': row_count,
        'sheet': data['sheet_name'],
    })

# ============================================================
# Шаг 3: Фильтруем и сортируем.
# ============================================================

# Только положительные долги
debtors = [r for r in results if r['debt'] > 0]
debtors.sort(key=lambda x: x['debt'], reverse=True)

# Клиенты с нулевым/отрицательным балансом
non_debtors = [r for r in results if r['debt'] <= 0]
non_debtors_negative = [r for r in results if r['debt'] < 0]

# ============================================================
# Шаг 4: Аналитика.
# ============================================================

total_debt = sum(d['debt'] for d in debtors)
total_rows_with_debt = sum(d['rows'] for d in debtors)
num_debtors = len(debtors)
avg_debt = total_debt / num_debtors if num_debtors > 0 else 0
max_debt = debtors[0]['debt'] if debtors else 0
max_debtor = debtors[0]['name'] if debtors else '-'
min_debt = debtors[-1]['debt'] if debtors else 0
min_debtor = debtors[-1]['name'] if debtors else '-'

# Распределение по листам
from collections import Counter
sheet_dist = Counter(d['sheet'] for d in debtors)

# ============================================================
# Шаг 5: Вывод отчёта.
# ============================================================

def fmt(n):
    """Форматирование числа с разделителями."""
    if n == int(n):
        return f"{int(n):,}".replace(',', ' ')
    return f"{n:,.2f}".replace(',', ' ')

print("=" * 90)
print("  ФИНАНСОВЫЙ АУДИТ ДОЛГОВ КЛИЕНТОВ")
print(f"  Источник: 03.03.2026.xlsx")
print(f"  Дата формирования: {__import__('datetime').datetime.now().strftime('%d.%m.%Y %H:%M')}")
print("=" * 90)

print(f"\n  Метод: колонка остатка ('Ост на...') из последнего листа, где клиент присутствует")
print(f"  Январь: кол. AA (27) | Февраль: кол. AB (28) | Март: кол. AB (28)")
print(f"  Не суммируются остатки разных месяцев.")

print(f"\n{'─' * 90}")
print(f"  СВОДКА")
print(f"{'─' * 90}")
print(f"  Общая сумма долга:          {fmt(total_debt)} сум")
print(f"  Количество должников:       {num_debtors}")
print(f"  Количество строк с долгом:  {total_rows_with_debt}")
print(f"  Средний долг клиента:       {fmt(avg_debt)} сум")
print(f"  Максимальный долг:          {fmt(max_debt)} сум ({max_debtor})")
print(f"  Минимальный долг:           {fmt(min_debt)} сум ({min_debtor})")

print(f"\n  Распределение должников по листам:")
for sheet, count in sorted(sheet_dist.items()):
    print(f"    {sheet}: {count} клиентов")

print(f"\n{'─' * 90}")
print(f"  ПОЛНЫЙ СПИСОК ДОЛЖНИКОВ ({num_debtors} клиентов)")
print(f"{'─' * 90}")

header = f"{'#':>4}  {'Клиент':<40} {'Долг (сум)':>20} {'Строк':>6}  {'Лист'}"
print(header)
print("─" * 90)

running_total = 0
for i, d in enumerate(debtors, 1):
    running_total += d['debt']
    name_display = d['name'][:38] + '..' if len(d['name']) > 40 else d['name']
    print(f"{i:>4}  {name_display:<40} {fmt(d['debt']):>20} {d['rows']:>6}  {d['sheet']}")

print("─" * 90)
print(f"{'':>4}  {'ИТОГО':<40} {fmt(total_debt):>20} {total_rows_with_debt:>6}")
print("=" * 90)

# =====ТОП-20 =====
print(f"\n{'─' * 90}")
print(f"  ТОП-20 КРУПНЕЙШИХ ДОЛЖНИКОВ")
print(f"{'─' * 90}")
print(f"{'#':>4}  {'Клиент':<40} {'Долг (сум)':>20} {'% от общего':>12}")
print("─" * 90)

for i, d in enumerate(debtors[:20], 1):
    pct = (d['debt'] / total_debt * 100) if total_debt > 0 else 0
    name_display = d['name'][:38] + '..' if len(d['name']) > 40 else d['name']
    print(f"{i:>4}  {name_display:<40} {fmt(d['debt']):>20} {pct:>11.1f}%")

top20_total = sum(d['debt'] for d in debtors[:20])
top20_pct = (top20_total / total_debt * 100) if total_debt > 0 else 0
print("─" * 90)
print(f"{'':>4}  {'Итого ТОП-20':<40} {fmt(top20_total):>20} {top20_pct:>11.1f}%")

# ===== Клиенты с переплатой =====
if non_debtors_negative:
    print(f"\n{'─' * 90}")
    print(f"  КЛИЕНТЫ С ПЕРЕПЛАТОЙ ({len(non_debtors_negative)})")
    print(f"{'─' * 90}")
    non_debtors_negative.sort(key=lambda x: x['debt'])
    for d in non_debtors_negative:
        print(f"  {d['name']:<40} {fmt(d['debt']):>20} сум  ({d['sheet']})")
    total_overpay = sum(d['debt'] for d in non_debtors_negative)
    print(f"  {'ИТОГО переплата':<40} {fmt(total_overpay):>20} сум")

print(f"\n{'=' * 90}")
print(f"  ФОРМУЛА РАСЧЁТА")
print(f"{'=' * 90}")
print(f"  1. Для каждого клиента берётся ПОСЛЕДНИЙ лист где он присутствует")
print(f"  2. Из этого листа берутся ВСЕ строки клиента")
print(f"  3. Debt(client) = SUM(balance_col) всех строк клиента в этом листе")
print(f"  4. balance_col = 'Ост на...' (AA/27 для янв, AB/28 для фев/мар)")
print(f"  5. Клиент считается должником если Debt(client) > 0")
print(f"  6. TotalDebt = SUM(Debt(client)) для всех должников")
print(f"\n  ВАЖНО: остатки разных месяцев НЕ суммируются.")
print(f"  Колонки M→AA (история платежей) НЕ используются для расчёта.")
print("=" * 90)
